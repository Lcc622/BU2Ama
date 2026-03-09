#!/usr/bin/env python3
"""加色加码 CLI 入口。"""

from __future__ import annotations

import argparse
import contextlib
import io
import sys
from pathlib import Path

from utils import bootstrap_app, normalize_store, print_result

bootstrap_app()

from app.config import RESULTS_DIR, STORE_CONFIGS, UPLOADS_DIR
import openpyxl


def count_template_rows(template_path: Path) -> int:
    """Count non-empty SKU rows in the template workbook."""
    workbook = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
        total = 0
        for row in worksheet.iter_rows(min_row=7, min_col=3, max_col=3, values_only=True):
            if row and row[0]:
                total += 1
        return total
    finally:
        workbook.close()


def resolve_source_files(store: str, provided_sources: list[str] | None) -> tuple[list[str], list[str]]:
    """Resolve source files for the selected store."""
    warnings: list[str] = []
    if provided_sources:
        source_files = [str(Path(name).name) for name in provided_sources if str(name).strip()]
    else:
        source_files = []
        config = STORE_CONFIGS.get({"EP": "EPUS", "DM": "DaMaUS", "PZ": "PZUS"}[store], {})
        for file_name in config.get("source_files", []):
            if (UPLOADS_DIR / file_name).exists():
                source_files.append(file_name)
        if not source_files:
            from app.core.excel_processor import ExcelProcessor
            processor = ExcelProcessor()
            scanned = processor._scan_store_data_files().get(store, [])  # noqa: SLF001
            source_files = list(scanned)

    existing: list[str] = []
    for source_file in source_files:
        if (UPLOADS_DIR / source_file).exists():
            existing.append(source_file)
        else:
            warnings.append(f"Source file not found and skipped: {source_file}")
    return existing, warnings


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="加色加码处理")
    parser.add_argument("--template", required=True, help="模板文件名或路径")
    parser.add_argument("--store", required=True, choices=["EP", "DM", "PZ"], help="店铺代码")
    parser.add_argument("--sources", nargs="+", help="源文件列表")
    parser.add_argument("--price-report", help="价格报告文件名或路径")
    parser.add_argument("--json", action="store_true", help="输出 JSON")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    result = {
        "success": False,
        "output_file": None,
        "processed_count": 0,
        "skipped_count": 0,
        "errors": [],
        "warnings": [],
    }

    try:
        store = normalize_store(args.store)
        template_name = Path(args.template).name
        template_path = UPLOADS_DIR / template_name
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_name}")

        source_files, warnings = resolve_source_files(store, args.sources)
        result["warnings"].extend(warnings)
        if not source_files:
            raise FileNotFoundError(f"No source files found for store {store}")

        price_report = Path(args.price_report).name if args.price_report else None
        if price_report and not (UPLOADS_DIR / price_report).exists():
            raise FileNotFoundError(f"Price report file not found: {price_report}")

        total_rows = count_template_rows(template_path)

        with contextlib.redirect_stdout(io.StringIO()):
            from app.core.excel_processor import ExcelProcessor
            processor = ExcelProcessor()
            output_filename, processed_count = processor.process_excel_new(
                template_filename=template_name,
                source_filenames=source_files,
                price_report_filename=price_report,
            )

        output_path = RESULTS_DIR / output_filename
        result.update(
            {
                "success": True,
                "output_file": str(output_path if output_path.exists() else output_filename),
                "processed_count": int(processed_count),
                "skipped_count": max(0, int(total_rows) - int(processed_count)),
            }
        )
        print_result(result, json_mode=args.json)
        return 0
    except Exception as exc:
        result["errors"].append(str(exc))
        print_result(result, json_mode=args.json)
        return 1


if __name__ == "__main__":
    sys.exit(main())
