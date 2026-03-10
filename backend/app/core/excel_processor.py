"""
Excel 处理模块
"""
import re
import csv
import pickle
import sqlite3
from copy import copy
from time import perf_counter
from threading import Lock
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from app.config import UPLOADS_DIR, RESULTS_DIR, TEMPLATES
from app.core.color_mapper import color_mapper
from app.models.excel import SKUInfo, ColorDistribution, AnalysisResult

STORE_PREFIX_PATTERN = re.compile(r"^(EP|DM|DA|PZ)-", re.IGNORECASE)
STORE_SOURCE_FILE_PATTERN = re.compile(r"^(EP|DM|DA|PZ)-\d", re.IGNORECASE)
STORE_PRIORITY = ("EP", "DM", "PZ")


class ExcelProcessor:
    """Excel 处理器"""

    def __init__(self):
        # 缓存：减少重复读取 EP 源表和价格报告
        self._price_map_cache: Dict[Tuple[str, Tuple[Tuple[str, int, int], ...]], Dict[str, float]] = {}
        self._asin_map_cache: Dict[Tuple[Tuple[str, int, int], ...], Dict[str, str]] = {}
        self._source_index_cache: Dict[Tuple[str, Tuple[Tuple[str, int, int], ...]], Dict[str, Any]] = {}
        # 缓存模板快照：避免每次任务都慢速 load_workbook(xlsm)
        self._template_snapshot_cache: Dict[Tuple[str, int, int], Dict[str, Any]] = {}
        self._template_cache_lock = Lock()
        legacy_db_path = UPLOADS_DIR / "excel_index.db"
        if legacy_db_path.exists():
            print("警告: 检测到旧索引库 excel_index.db，请重新上传店铺数据源以重建分店铺索引。")

    def get_store_prefix(self, filename: str) -> str:
        match = STORE_PREFIX_PATTERN.match(str(filename or "").strip().upper())
        if match:
            return self._normalize_store_prefix(match.group(1))
        return "EP"

    def _normalize_store_prefix(self, store_prefix: Optional[str]) -> str:
        normalized = str(store_prefix or "").strip().upper()
        if normalized == "DA":
            normalized = "DM"
        if normalized not in STORE_PRIORITY:
            normalized = "EP"
        return normalized

    def get_index_db_path(self, store_prefix: str) -> Path:
        normalized = self._normalize_store_prefix(store_prefix)
        return UPLOADS_DIR / f"excel_index_{normalized}.db"

    def _resolve_store_for_filename(self, filename: Optional[str], fallback_store: Optional[str] = None) -> str:
        fallback = self._normalize_store_prefix(fallback_store)
        if not filename:
            return fallback
        if STORE_PREFIX_PATTERN.match(str(filename).strip().upper()):
            return self.get_store_prefix(filename)
        return fallback

    def _guess_store_from_template_type(self, template_type: Optional[str]) -> Optional[str]:
        text = str(template_type or "").strip().upper()
        if not text:
            return None
        if "PZ" in text:
            return "PZ"
        if "DM" in text or "DAMA" in text or text.startswith("DA"):
            return "DM"
        if "EP" in text:
            return "EP"
        return None

    def _get_store_suffix_family(self, template_type: Optional[str]) -> Tuple[str, str]:
        """返回当前模板所属店铺的标准后缀族。

        Returns:
            (base_suffix, plus_suffix)

        约定：
        - EP: ``-USA`` / ``-PH``
        - DAMA(DM): ``-PL`` / ``-PLPH``
        - PZ: ``-DA`` / ``-DAPH``
        """
        store_prefix = self._guess_store_from_template_type(template_type)
        normalized_store = self._normalize_store_prefix(store_prefix)
        if normalized_store == "DM":
            return "-PL", "-PLPH"
        if normalized_store == "PZ":
            return "-DA", "-DAPH"
        return "-USA", "-PH"

    def _is_plus_body_suffix(self, suffix: Optional[str]) -> bool:
        """判断后缀是否属于大码/Plus Body 后缀。"""
        normalized = str(suffix or "").strip().upper()
        if not normalized:
            return False
        if not normalized.startswith("-"):
            normalized = f"-{normalized}"
        return normalized in {"-PH", "-PLPH", "-DAPH"}

    def _generate_suffix(self, template_type: Optional[str], size: Any) -> str:
        """根据模板类型和尺码生成标准后缀。"""
        size_text = str(size or "")
        match = re.search(r"\d+", size_text)
        size_num = int(match.group()) if match else 0
        base_suffix, plus_suffix = self._get_store_suffix_family(template_type)
        return plus_suffix if size_num >= 14 else base_suffix

    def _guess_store_from_skus(self, skus: Optional[List[str]]) -> Optional[str]:
        if not skus:
            return None
        for sku in skus:
            match = STORE_PREFIX_PATTERN.match(str(sku or "").strip().upper())
            if match:
                return self._normalize_store_prefix(match.group(1))
        return None

    def _build_store_search_order(self, preferred_store: Optional[str]) -> List[str]:
        normalized = self._normalize_store_prefix(preferred_store)
        if normalized in STORE_PRIORITY:
            return [normalized] + [store for store in STORE_PRIORITY if store != normalized]
        return list(STORE_PRIORITY)

    def _scan_store_data_files(self) -> Dict[str, List[str]]:
        files_by_store: Dict[str, List[str]] = {store: [] for store in STORE_PRIORITY}
        for path in sorted(UPLOADS_DIR.iterdir()):
            if not path.is_file():
                continue
            if path.suffix.lower() not in {".xlsm", ".xlsx"}:
                continue
            if not STORE_PREFIX_PATTERN.match(path.name.strip().upper()):
                continue
            if not STORE_SOURCE_FILE_PATTERN.match(path.name.strip().upper()):
                continue
            store_prefix = self.get_store_prefix(path.name)
            files_by_store[store_prefix].append(path.name)
        return files_by_store

    def _build_store_data_files(
        self,
        request_data_files: List[str],
        context_store: Optional[str],
    ) -> Dict[str, List[str]]:
        scanned_files = self._scan_store_data_files()
        normalized_context = self._normalize_store_prefix(context_store)
        for file_name in request_data_files:
            store_prefix = self._resolve_store_for_filename(file_name, normalized_context)
            if file_name not in scanned_files[store_prefix]:
                scanned_files[store_prefix].append(file_name)
        return {
            store: sorted({name for name in names})
            for store, names in scanned_files.items()
        }

    def _signature_to_key(self, signature: Tuple[Tuple[str, int, int], ...]) -> str:
        return "|".join([f"{name}:{mtime}:{size}" for name, mtime, size in signature])

    def _connect_index_db(self, store_prefix: str) -> sqlite3.Connection:
        conn = sqlite3.connect(self.get_index_db_path(store_prefix))
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        return conn

    def _ensure_index_tables(self, conn: sqlite3.Connection) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS price_index (
                signature TEXT NOT NULL,
                sku TEXT NOT NULL,
                price REAL NOT NULL,
                PRIMARY KEY (signature, sku)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_price_signature ON price_index(signature)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS source_rows_index (
                signature TEXT NOT NULL,
                sku TEXT NOT NULL,
                sku_base TEXT NOT NULL,
                product_code TEXT NOT NULL,
                color_code TEXT NOT NULL,
                size TEXT NOT NULL,
                suffix TEXT NOT NULL,
                data_file TEXT NOT NULL,
                row_blob BLOB NOT NULL,
                PRIMARY KEY (signature, sku)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_source_signature ON source_rows_index(signature)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_source_style ON source_rows_index(signature, product_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_source_style_size ON source_rows_index(signature, product_code, size)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_source_style_color ON source_rows_index(signature, product_code, color_code)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS source_headers_index (
                signature TEXT NOT NULL,
                data_file TEXT NOT NULL,
                header_norm TEXT NOT NULL,
                col_idx INTEGER NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_headers_signature ON source_headers_index(signature, data_file)")
        conn.commit()

    def _load_price_map_from_sqlite(self, signature_key: str, store_prefix: str) -> Optional[Dict[str, float]]:
        conn = self._connect_index_db(store_prefix)
        try:
            self._ensure_index_tables(conn)
            rows = conn.execute(
                "SELECT sku, price FROM price_index WHERE signature = ?",
                (signature_key,)
            ).fetchall()
            if not rows:
                return None
            return {str(sku): float(price) for sku, price in rows}
        finally:
            conn.close()

    def _save_price_map_to_sqlite(self, signature_key: str, price_map: Dict[str, float], store_prefix: str) -> None:
        conn = self._connect_index_db(store_prefix)
        try:
            self._ensure_index_tables(conn)
            conn.execute("DELETE FROM price_index WHERE signature = ?", (signature_key,))
            if price_map:
                conn.executemany(
                    "INSERT INTO price_index(signature, sku, price) VALUES (?, ?, ?)",
                    [(signature_key, sku, float(price)) for sku, price in price_map.items()]
                )
            conn.commit()
        finally:
            conn.close()

    def _load_source_index_from_sqlite(self, signature_key: str, store_prefix: str) -> Optional[Dict[str, Any]]:
        conn = self._connect_index_db(store_prefix)
        try:
            self._ensure_index_tables(conn)
            source_rows_db = conn.execute(
                """
                SELECT sku, sku_base, product_code, color_code, size, suffix, data_file, row_blob
                FROM source_rows_index
                WHERE signature = ?
                """,
                (signature_key,)
            ).fetchall()
            if not source_rows_db:
                return None

            header_rows = conn.execute(
                """
                SELECT data_file, header_norm, col_idx
                FROM source_headers_index
                WHERE signature = ?
                ORDER BY data_file, header_norm, col_idx
                """,
                (signature_key,)
            ).fetchall()
        finally:
            conn.close()

        sku_to_source: Dict[str, str] = {}
        sku_base_to_source: Dict[str, str] = {}
        style_suffix_to_source: Dict[Tuple[str, str], str] = {}
        style_to_source: Dict[str, str] = {}
        style_size_to_source: Dict[Tuple[str, str], str] = {}
        style_size_suffix_to_source: Dict[Tuple[str, str, str], str] = {}
        style_color_to_source: Dict[Tuple[str, str], str] = {}
        style_color_suffix_to_source: Dict[Tuple[str, str, str], str] = {}
        source_rows: Dict[str, List[Any]] = {}
        source_file_by_sku: Dict[str, str] = {}
        source_header_map_by_file: Dict[str, Dict[str, List[int]]] = {}

        def prefer_size_02(existing_sku: Optional[str], candidate_sku: str) -> str:
            if existing_sku is None:
                return candidate_sku
            ex_info = self.parse_sku(existing_sku)
            cand_info = self.parse_sku(candidate_sku)
            if ex_info and cand_info and ex_info.size != "02" and cand_info.size == "02":
                return candidate_sku
            return existing_sku

        for sku, sku_base, product_code, color_code, size, suffix, data_file, row_blob in source_rows_db:
            sku_str = str(sku)
            sku_to_source[sku_str] = sku_str
            sku_base_to_source[str(sku_base)] = sku_str
            source_rows[sku_str] = pickle.loads(row_blob)
            source_file_by_sku[sku_str] = str(data_file)

            parsed = self.parse_sku(sku_str)
            if not parsed:
                continue

            style_key = parsed.product_code
            suffix_key = parsed.suffix
            size_key = parsed.size
            color_key = parsed.color_code

            style_suffix_to_source[(style_key, suffix_key)] = prefer_size_02(
                style_suffix_to_source.get((style_key, suffix_key)),
                sku_str
            )
            style_to_source[style_key] = prefer_size_02(style_to_source.get(style_key), sku_str)
            style_size_to_source.setdefault((style_key, size_key), sku_str)
            style_size_suffix_to_source.setdefault((style_key, size_key, suffix_key), sku_str)
            style_color_to_source[(style_key, color_key)] = prefer_size_02(
                style_color_to_source.get((style_key, color_key)),
                sku_str
            )
            style_color_suffix_to_source[(style_key, color_key, suffix_key)] = prefer_size_02(
                style_color_suffix_to_source.get((style_key, color_key, suffix_key)),
                sku_str
            )

        for data_file, header_norm, col_idx in header_rows:
            file_key = str(data_file)
            if file_key not in source_header_map_by_file:
                source_header_map_by_file[file_key] = {}
            header_key = str(header_norm)
            source_header_map_by_file[file_key].setdefault(header_key, []).append(int(col_idx))

        return {
            "sku_to_source": sku_to_source,
            "sku_base_to_source": sku_base_to_source,
            "style_suffix_to_source": style_suffix_to_source,
            "style_to_source": style_to_source,
            "style_size_to_source": style_size_to_source,
            "style_size_suffix_to_source": style_size_suffix_to_source,
            "style_color_to_source": style_color_to_source,
            "style_color_suffix_to_source": style_color_suffix_to_source,
            "source_rows": source_rows,
            "source_file_by_sku": source_file_by_sku,
            "source_header_map_by_file": source_header_map_by_file,
        }

    def _save_source_index_to_sqlite(self, signature_key: str, index_data: Dict[str, Any], store_prefix: str) -> None:
        conn = self._connect_index_db(store_prefix)
        try:
            self._ensure_index_tables(conn)
            conn.execute("DELETE FROM source_rows_index WHERE signature = ?", (signature_key,))
            conn.execute("DELETE FROM source_headers_index WHERE signature = ?", (signature_key,))

            sku_to_source: Dict[str, str] = index_data["sku_to_source"]
            source_rows: Dict[str, List[Any]] = index_data["source_rows"]
            source_file_by_sku: Dict[str, str] = index_data["source_file_by_sku"]
            source_header_map_by_file: Dict[str, Dict[str, List[int]]] = index_data["source_header_map_by_file"]

            row_batch: List[Tuple[str, str, str, str, str, str, str, str, bytes]] = []
            for source_sku in sku_to_source.keys():
                parsed = self.parse_sku(source_sku)
                product_code = parsed.product_code if parsed else (source_sku[:7] if len(source_sku) >= 7 else "")
                color_code = parsed.color_code if parsed else (source_sku[7:9] if len(source_sku) >= 9 else "")
                size = parsed.size if parsed else (source_sku[9:11] if len(source_sku) >= 11 else "")
                suffix = parsed.suffix if parsed else (source_sku[11:] if len(source_sku) > 11 else "")
                sku_base = source_sku[:11] if len(source_sku) >= 11 else source_sku
                row_values = source_rows.get(source_sku, [])
                row_blob = pickle.dumps(row_values, protocol=pickle.HIGHEST_PROTOCOL)
                row_batch.append(
                    (
                        signature_key,
                        source_sku,
                        sku_base,
                        product_code,
                        color_code,
                        size,
                        suffix,
                        source_file_by_sku.get(source_sku, ""),
                        row_blob,
                    )
                )

            if row_batch:
                conn.executemany(
                    """
                    INSERT INTO source_rows_index(
                        signature, sku, sku_base, product_code, color_code, size, suffix, data_file, row_blob
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    row_batch
                )

            header_batch: List[Tuple[str, str, str, int]] = []
            for data_file, header_map in source_header_map_by_file.items():
                for header_norm, cols in header_map.items():
                    for col_idx in cols:
                        header_batch.append((signature_key, data_file, header_norm, int(col_idx)))
            if header_batch:
                conn.executemany(
                    "INSERT INTO source_headers_index(signature, data_file, header_norm, col_idx) VALUES (?, ?, ?, ?)",
                    header_batch
                )
            conn.commit()
        finally:
            conn.close()

    def _normalize_header(self, value: str) -> str:
        return str(value).strip().lower()

    def _detect_effective_max_col(self, ws: Worksheet, header_rows: List[int], hard_cap: Optional[int] = None) -> int:
        max_col = ws.max_column
        if hard_cap is not None:
            max_col = min(max_col, hard_cap)
        effective = 1
        for row_idx in header_rows:
            if row_idx > ws.max_row:
                continue
            for col_idx in range(max_col, 0, -1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None and str(value).strip() != "":
                    if col_idx > effective:
                        effective = col_idx
                    break
        return effective

    def _build_files_signature(self, filenames: List[str]) -> Tuple[Tuple[str, int, int], ...]:
        signature: List[Tuple[str, int, int]] = []
        for filename in sorted(filenames):
            path = UPLOADS_DIR / filename
            if not path.exists():
                continue
            stat = path.stat()
            signature.append((filename, int(stat.st_mtime_ns), int(stat.st_size)))
        return tuple(signature)

    def _build_template_signature(self, template_file: str) -> Tuple[str, int, int]:
        path = UPLOADS_DIR / template_file
        stat = path.stat()
        return (template_file, int(stat.st_mtime_ns), int(stat.st_size))

    def _load_template_snapshot_cached(self, template_file: str) -> Tuple[Dict[str, Any], bool]:
        signature = self._build_template_signature(template_file)
        with self._template_cache_lock:
            cached = self._template_snapshot_cache.get(signature)
            if cached is not None:
                return cached, True

        template_path = UPLOADS_DIR / template_file
        template_wb = openpyxl.load_workbook(template_path)
        try:
            template_ws = template_wb['Template'] if 'Template' in template_wb.sheetnames else template_wb.active
            effective_max_col = self._detect_effective_max_col(template_ws, [2, 3, 4, 5], hard_cap=700)

            def snapshot_row(row_idx: int) -> Tuple[List[Any], List[Dict[str, Any]], Optional[float]]:
                values: List[Any] = []
                styles: List[Dict[str, Any]] = []
                for col_idx in range(1, effective_max_col + 1):
                    cell = template_ws.cell(row=row_idx, column=col_idx)
                    values.append(cell.value)
                    styles.append({
                        "font": copy(cell.font),
                        "border": copy(cell.border),
                        "fill": copy(cell.fill),
                        "number_format": cell.number_format,
                        "protection": copy(cell.protection),
                        "alignment": copy(cell.alignment),
                    })
                row_height = template_ws.row_dimensions[row_idx].height if row_idx in template_ws.row_dimensions else None
                return values, styles, row_height

            row_snapshots = {
                row_idx: snapshot_row(row_idx)
                for row_idx in (1, 2, 3, 4, 5)
            }
            column_widths: Dict[int, Optional[float]] = {}
            for col_idx in range(1, effective_max_col + 1):
                col_letter = get_column_letter(col_idx)
                width = template_ws.column_dimensions[col_letter].width if col_letter in template_ws.column_dimensions else None
                column_widths[col_idx] = width

            snapshot = {
                "effective_max_col": effective_max_col,
                "row_snapshots": row_snapshots,
                "column_widths": column_widths,
            }
        finally:
            template_wb.close()

        with self._template_cache_lock:
            # 清理同名旧版本快照，避免模板更新后缓存累积
            old_keys = [k for k in self._template_snapshot_cache.keys() if k[0] == template_file and k != signature]
            for key in old_keys:
                self._template_snapshot_cache.pop(key, None)
            self._template_snapshot_cache[signature] = snapshot
        return snapshot, False

    def prewarm_template_cache(self, template_files: Optional[List[str]] = None) -> Dict[str, str]:
        """预热模板缓存，让首次处理也能命中内存缓存。"""
        files = template_files or [cfg["template_file"] for cfg in TEMPLATES.values()]
        results: Dict[str, str] = {}
        for template_file in files:
            template_path = UPLOADS_DIR / template_file
            if not template_path.exists():
                results[template_file] = "missing"
                continue
            try:
                _, cache_hit = self._load_template_snapshot_cached(template_file)
                results[template_file] = "hit" if cache_hit else "loaded"
            except Exception:
                results[template_file] = "error"
        return results

    def _load_price_map_cached(self, price_report_file: Optional[str], store_prefix: str) -> Dict[str, float]:
        if not price_report_file:
            return {}
        normalized_store = self._resolve_store_for_filename(price_report_file, store_prefix)
        signature = self._build_files_signature([price_report_file])
        signature_key = self._signature_to_key(signature)
        cache_key = (normalized_store, signature)
        if cache_key in self._price_map_cache:
            print(f"价格映射命中缓存 {len(self._price_map_cache[cache_key])} 条")
            return self._price_map_cache[cache_key]
        sqlite_cached = self._load_price_map_from_sqlite(signature_key, normalized_store)
        if sqlite_cached is not None:
            self._price_map_cache[cache_key] = sqlite_cached
            print(f"价格映射命中 SQLite 缓存 {len(sqlite_cached)} 条")
            return sqlite_cached

        price_map: Dict[str, float] = {}
        price_path = UPLOADS_DIR / price_report_file
        if price_path.exists():
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    with open(price_path, 'r', encoding=encoding) as f:
                        reader = csv.DictReader(f, delimiter='\t')
                        for row in reader:
                            sku = str(row.get('seller-sku', '')).strip().upper()
                            price_text = str(row.get('price', '')).strip()
                            if not sku or not price_text:
                                continue
                            try:
                                price_map[sku] = float(price_text)
                            except (ValueError, TypeError):
                                pass
                    break
                except UnicodeDecodeError:
                    continue
        self._price_map_cache[cache_key] = price_map
        self._save_price_map_to_sqlite(signature_key, price_map, normalized_store)
        print(f"价格映射加载 {len(price_map)} 条")
        return price_map

    def _load_asin_map_cached(self, price_report_file: Optional[str]) -> Dict[str, str]:
        if not price_report_file:
            return {}
        signature = self._build_files_signature([price_report_file])
        if signature in self._asin_map_cache:
            print(f"ASIN 映射命中缓存 {len(self._asin_map_cache[signature])} 条")
            return self._asin_map_cache[signature]

        asin_map: Dict[str, str] = {}
        price_path = UPLOADS_DIR / price_report_file
        if price_path.exists():
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    with open(price_path, 'r', encoding=encoding) as f:
                        reader = csv.DictReader(f, delimiter='\t')
                        for row in reader:
                            sku = str(row.get('seller-sku', '')).strip().upper()
                            asin = str(row.get('asin1', '')).strip().upper()
                            if not sku or not asin:
                                continue
                            asin_map[sku] = asin
                    break
                except UnicodeDecodeError:
                    continue
        self._asin_map_cache[signature] = asin_map
        print(f"ASIN 映射加载 {len(asin_map)} 条")
        return asin_map

    def _build_source_index_cached(self, data_files: List[str], store_prefix: str) -> Dict[str, Any]:
        normalized_store = self._resolve_store_for_filename(None, store_prefix)
        signature = self._build_files_signature(data_files)
        signature_key = self._signature_to_key(signature)
        cache_key = (normalized_store, signature)
        if cache_key in self._source_index_cache:
            cached = self._source_index_cache[cache_key]
            print(
                f"[{normalized_store}] 数据索引命中缓存 sku={len(cached['sku_to_source'])}, "
                f"style={len(cached['style_to_source'])}"
            )
            return cached
        sqlite_cached = self._load_source_index_from_sqlite(signature_key, normalized_store)
        if sqlite_cached is not None:
            self._source_index_cache[cache_key] = sqlite_cached
            print(
                f"[{normalized_store}] 数据索引命中 SQLite 缓存 sku={len(sqlite_cached['sku_to_source'])}, "
                f"style={len(sqlite_cached['style_to_source'])}"
            )
            return sqlite_cached

        sku_to_source: Dict[str, str] = {}
        sku_base_to_source: Dict[str, str] = {}
        style_suffix_to_source: Dict[Tuple[str, str], str] = {}
        style_to_source: Dict[str, str] = {}
        style_size_to_source: Dict[Tuple[str, str], str] = {}
        style_size_suffix_to_source: Dict[Tuple[str, str, str], str] = {}
        style_color_to_source: Dict[Tuple[str, str], str] = {}
        style_color_suffix_to_source: Dict[Tuple[str, str, str], str] = {}
        source_rows: Dict[str, List[Any]] = {}
        source_file_by_sku: Dict[str, str] = {}
        source_header_map_by_file: Dict[str, Dict[str, List[int]]] = {}

        for data_file in data_files:
            data_path = UPLOADS_DIR / data_file
            if not data_path.exists():
                print(f"数据文件不存在，跳过: {data_file}")
                continue

            # 只读模式更快，且仅用于读取源数据
            data_wb = openpyxl.load_workbook(data_path, data_only=True, read_only=True)
            data_ws = data_wb['Template'] if 'Template' in data_wb.sheetnames else data_wb.active
            source_effective_max_col = self._detect_effective_max_col(data_ws, [4], hard_cap=800)

            header_sku_col = 3 if str(data_ws.cell(row=4, column=3).value or '').strip().upper() == 'SKU' else 2
            source_header_map: Dict[str, List[int]] = defaultdict(list)
            for col_idx in range(1, source_effective_max_col + 1):
                source_header = data_ws.cell(row=4, column=col_idx).value
                if source_header:
                    source_header_map[self._normalize_header(source_header)].append(col_idx)
            source_header_map_by_file[data_file] = dict(source_header_map)

            for row in data_ws.iter_rows(min_row=5, max_row=data_ws.max_row, min_col=1, max_col=source_effective_max_col, values_only=True):
                sku_value = row[header_sku_col - 1] if len(row) >= header_sku_col else None
                if not sku_value:
                    continue
                source_sku = str(sku_value).strip().upper()
                if not source_sku or source_sku == 'SKU':
                    continue

                row_values = list(row)
                if source_sku not in sku_to_source:
                    sku_to_source[source_sku] = source_sku
                    source_rows[source_sku] = row_values
                    source_file_by_sku[source_sku] = data_file
                if len(source_sku) >= 11:
                    sku_base = source_sku[:11]
                    if sku_base not in sku_base_to_source:
                        sku_base_to_source[sku_base] = source_sku

                source_info = self.parse_sku(source_sku)
                if source_info:
                    style_key = source_info.product_code
                    suffix_key = source_info.suffix
                    style_suffix_key = (style_key, suffix_key)

                    existing_sku = style_suffix_to_source.get(style_suffix_key)
                    if existing_sku is None:
                        style_suffix_to_source[style_suffix_key] = source_sku
                    else:
                        ex_info = self.parse_sku(existing_sku)
                        if ex_info and ex_info.size != "02" and source_info.size == "02":
                            style_suffix_to_source[style_suffix_key] = source_sku

                    existing_style_sku = style_to_source.get(style_key)
                    if existing_style_sku is None:
                        style_to_source[style_key] = source_sku
                    else:
                        ex_info2 = self.parse_sku(existing_style_sku)
                        if ex_info2 and ex_info2.size != "02" and source_info.size == "02":
                            style_to_source[style_key] = source_sku

                    size_key = (style_key, source_info.size)
                    if size_key not in style_size_to_source:
                        style_size_to_source[size_key] = source_sku
                    suffix_size_key = (style_key, source_info.size, suffix_key)
                    if suffix_size_key not in style_size_suffix_to_source:
                        style_size_suffix_to_source[suffix_size_key] = source_sku

                    color_key = (style_key, source_info.color_code)
                    existing_color_sku = style_color_to_source.get(color_key)
                    if existing_color_sku is None:
                        style_color_to_source[color_key] = source_sku
                    else:
                        ex_color_info = self.parse_sku(existing_color_sku)
                        if ex_color_info and ex_color_info.size != "02" and source_info.size == "02":
                            style_color_to_source[color_key] = source_sku

                    suffix_color_key = (style_key, source_info.color_code, suffix_key)
                    existing_suffix_color_sku = style_color_suffix_to_source.get(suffix_color_key)
                    if existing_suffix_color_sku is None:
                        style_color_suffix_to_source[suffix_color_key] = source_sku
                    else:
                        ex_suffix_color_info = self.parse_sku(existing_suffix_color_sku)
                        if ex_suffix_color_info and ex_suffix_color_info.size != "02" and source_info.size == "02":
                            style_color_suffix_to_source[suffix_color_key] = source_sku

            data_wb.close()

        index_data: Dict[str, Any] = {
            "sku_to_source": sku_to_source,
            "sku_base_to_source": sku_base_to_source,
            "style_suffix_to_source": style_suffix_to_source,
            "style_to_source": style_to_source,
            "style_size_to_source": style_size_to_source,
            "style_size_suffix_to_source": style_size_suffix_to_source,
            "style_color_to_source": style_color_to_source,
            "style_color_suffix_to_source": style_color_suffix_to_source,
            "source_rows": source_rows,
            "source_file_by_sku": source_file_by_sku,
            "source_header_map_by_file": source_header_map_by_file,
        }
        self._source_index_cache[cache_key] = index_data
        self._save_source_index_to_sqlite(signature_key, index_data, normalized_store)
        print(f"[{normalized_store}] 数据索引加载 sku={len(sku_to_source)}, style={len(style_to_source)}")
        return index_data

    def _build_source_indexes_cached(self, data_files: List[str], context_store: Optional[str]) -> Dict[str, Dict[str, Any]]:
        files_by_store = self._build_store_data_files(data_files, context_store)
        has_explicit_store = any(
            STORE_PREFIX_PATTERN.match(str(file_name).strip().upper())
            for file_name in data_files
        )
        normalized_context_store = self._normalize_store_prefix(context_store)
        if context_store is not None or has_explicit_store:
            requested_store_set = {
                self._resolve_store_for_filename(file_name, normalized_context_store)
                for file_name in data_files
            }
            if not requested_store_set:
                fallback_store = normalized_context_store
                requested_store_set = {fallback_store}
        else:
            requested_store_set = set(STORE_PRIORITY)

        indexes_by_store: Dict[str, Dict[str, Any]] = {}
        for store_prefix in STORE_PRIORITY:
            if store_prefix not in requested_store_set:
                continue
            store_files = files_by_store.get(store_prefix, [])
            if not store_files:
                continue
            indexes_by_store[store_prefix] = self._build_source_index_cached(store_files, store_prefix)
        return indexes_by_store

    def _merge_source_indexes(
        self,
        indexes_by_store: Dict[str, Dict[str, Any]],
        store_order: List[str],
    ) -> Dict[str, Any]:
        keys = [
            "sku_to_source",
            "sku_base_to_source",
            "style_suffix_to_source",
            "style_to_source",
            "style_size_to_source",
            "style_size_suffix_to_source",
            "style_color_to_source",
            "style_color_suffix_to_source",
            "source_rows",
            "source_file_by_sku",
        ]
        merged: Dict[str, Any] = {key: {} for key in keys}
        merged["source_header_map_by_file"] = {}

        for store_prefix in store_order:
            index_data = indexes_by_store.get(store_prefix)
            if not index_data:
                continue
            for key in keys:
                dest = merged[key]
                for data_key, value in index_data[key].items():
                    if data_key not in dest:
                        dest[data_key] = value
            source_header_map = merged["source_header_map_by_file"]
            for data_file, header_map in index_data["source_header_map_by_file"].items():
                if data_file not in source_header_map:
                    source_header_map[data_file] = header_map

        return merged

    def parse_sku(self, sku: str) -> Optional[SKUInfo]:
        """解析 SKU 格式

        格式：前7位(Style) + 2位(颜色) + 2位(尺码) + 后缀
        例如：ES0128BDG02-PH
        - ES0128B: Style (7位)
        - DG: 颜色代码 (2位)
        - 02: 尺码 (2位)
        - -PH: 后缀
        """
        sku = sku.strip().upper()

        # 最小长度：7(style) + 2(color) + 2(size) = 11
        if len(sku) < 11:
            return None

        # 按位置提取
        product_code = sku[:7]  # 前7位
        color_code = sku[7:9]   # 8-9位
        size = sku[9:11]        # 10-11位
        suffix = sku[11:] if len(sku) > 11 else ""  # 剩余部分

        # 验证格式
        if not color_code.isalpha() or not color_code.isupper():
            return None
        if not size.isdigit():
            return None

        return SKUInfo(
            sku=sku,
            product_code=product_code,
            color_code=color_code,
            size=size,
            suffix=suffix
        )

    def extract_color_from_sku(self, sku: str) -> Optional[str]:
        """从 SKU 中提取颜色代码"""
        info = self.parse_sku(sku)
        return info.color_code if info else None

    def read_skus_from_txt(self, filename: str) -> List[str]:
        """从 .txt 文件中读取 SKU 列表

        支持 TSV 格式（Tab-Separated Values），SKU 在 seller-sku 列
        """
        filepath = UPLOADS_DIR / filename

        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filename}")

        skus = []

        # 尝试多种编码
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']

        for encoding in encodings:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    reader = csv.DictReader(f, delimiter='\t')
                    for row in reader:
                        sku = row.get('seller-sku', '').strip()
                        if sku:
                            skus.append(sku)
                print(f"成功使用 {encoding} 编码读取文件")
                break
            except (UnicodeDecodeError, Exception) as e:
                if encoding == encodings[-1]:
                    # 最后一个编码也失败了
                    raise Exception(f"无法读取文件 {filename}，尝试了所有编码: {encodings}")
                continue

        return skus

    def read_skus_from_excel(self, filename: str) -> List[str]:
        """从 Excel 文件中读取 SKU 列表"""
        filepath = UPLOADS_DIR / filename

        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filename}")

        wb = openpyxl.load_workbook(filepath, data_only=True)

        if 'Template' in wb.sheetnames:
            ws = wb['Template']
        else:
            ws = wb.active

        skus = []

        # 检测文件格式
        row4 = list(ws[4])
        if len(row4) > 2 and row4[2].value and 'SKU' in str(row4[2].value):
            # EP-0/1/2.xlsm 格式
            sku_col = 2
            start_row = 7
        else:
            # 其他模板格式
            sku_col = 1
            start_row = 4

        for row_idx in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=sku_col)
            if cell.value:
                sku = str(cell.value).strip()
                if sku:
                    skus.append(sku)

        return skus

    def analyze_excel_file(self, filename: str) -> AnalysisResult:
        """分析 Excel 文件"""
        filepath = UPLOADS_DIR / filename

        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filename}")

        # 读取 Excel 文件
        wb = openpyxl.load_workbook(filepath, data_only=True)

        # 尝试使用 Template 工作表，如果不存在则使用活动工作表
        if 'Template' in wb.sheetnames:
            ws = wb['Template']
        else:
            ws = wb.active

        # 统计数据
        color_counts: Dict[str, int] = {}
        unknown_colors: set = set()
        prefixes: set = set()
        suffixes: set = set()
        total_skus = 0

        # 检测文件格式
        # EP-0/1/2.xlsm: SKU 在列 2（索引 2），从第 7 行开始
        # 其他模板: SKU 在列 1（索引 1），从第 4 行开始

        # 检查第 4 行的字段名
        row4 = list(ws[4])
        if len(row4) > 2 and row4[2].value and 'SKU' in str(row4[2].value):
            # EP-0/1/2.xlsm 格式
            sku_col = 2
            start_row = 7
        else:
            # 其他模板格式
            sku_col = 1
            start_row = 4

        # 遍历所有行
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if not row or len(row) <= sku_col or not row[sku_col]:
                continue

            sku = str(row[sku_col]).strip()
            info = self.parse_sku(sku)

            if info:
                total_skus += 1

                # 统计颜色
                color_code = info.color_code
                color_counts[color_code] = color_counts.get(color_code, 0) + 1

                # 检查是否为未知颜色
                if not color_mapper.get_color_name(color_code):
                    unknown_colors.add(color_code)

                # 收集前缀和后缀
                prefixes.add(info.product_code)
                if info.suffix:
                    suffixes.add(info.suffix)

        wb.close()

        # 构建颜色分布列表
        color_distribution = [
            ColorDistribution(
                color_code=code,
                color_name=color_mapper.get_color_name(code),
                count=count
            )
            for code, count in sorted(color_counts.items(), key=lambda x: x[1], reverse=True)
        ]

        return AnalysisResult(
            success=True,
            filename=filename,
            total_skus=total_skus,
            unique_colors=len(color_counts),
            color_distribution=color_distribution,
            unknown_colors=sorted(unknown_colors),
            prefixes=sorted(prefixes),
            suffixes=sorted(suffixes)
        )

    def get_color_map_value(self, color_name: str) -> str:
        """获取颜色分类"""
        color_lower = color_name.lower()

        color_categories = {
            'Purple': ['purple', 'violet', 'lavender', 'plum'],
            'Blue': ['blue', 'navy', 'azure', 'cyan', 'teal'],
            'Green': ['green', 'olive', 'lime', 'mint'],
            'Red': ['red', 'crimson', 'burgundy', 'wine'],
            'Pink': ['pink', 'rose', 'coral', 'fuchsia'],
            'Orange': ['orange', 'peach', 'apricot'],
            'Yellow': ['yellow', 'gold', 'cream', 'beige'],
            'Brown': ['brown', 'tan', 'khaki', 'coffee', 'chocolate'],
            'Black': ['black'],
            'White': ['white', 'ivory'],
            'Grey': ['grey', 'gray', 'silver'],
            'Multicolor': ['multicolor', 'multi', 'print', 'pattern']
        }

        for category, keywords in color_categories.items():
            if any(keyword in color_lower for keyword in keywords):
                return category

        return 'Multicolor'

    def calculate_launch_date(self) -> str:
        """计算上线日期（北京时间逻辑）"""
        # 获取当前 UTC 时间
        utc_now = datetime.utcnow()

        # 转换为北京时间（UTC+8）
        beijing_time = utc_now + timedelta(hours=8)

        # 如果北京时间在下午 3 点之前，使用前一天
        if beijing_time.hour < 15:
            launch_date = beijing_time - timedelta(days=1)
        else:
            launch_date = beijing_time

        return launch_date.strftime('%Y-%m-%d')

    def process_excel_new(
        self,
        template_filename: str,
        source_filenames: List[str],
        price_report_filename: Optional[str] = None
    ) -> Tuple[str, int]:
        """新的处理逻辑：从源文件中查找 SKU 数据并更新模板

        Args:
            template_filename: 输入模板文件（包含需要处理的 SKU）
            source_filenames: 数据源文件列表（EP-0/1/2.xlsm）
            price_report_filename: 价格报告文件（All Listings Report）

        Returns:
            (输出文件名, 处理的行数)
        """

        # 1. 读取价格报告
        price_map = {}
        if price_report_filename:
            price_file = UPLOADS_DIR / price_report_filename
            if price_file.exists():
                # 尝试不同的编码
                for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                    try:
                        with open(price_file, 'r', encoding=encoding) as f:
                            lines = f.readlines()
                            if len(lines) > 1:
                                # 第一行是表头
                                headers = lines[0].strip().split('\t')
                                sku_idx = headers.index('seller-sku') if 'seller-sku' in headers else -1
                                price_idx = headers.index('price') if 'price' in headers else -1

                                if sku_idx >= 0 and price_idx >= 0:
                                    for line in lines[1:]:
                                        fields = line.strip().split('\t')
                                        if len(fields) > max(sku_idx, price_idx):
                                            sku = fields[sku_idx]
                                            try:
                                                price = float(fields[price_idx])
                                                price_map[sku] = price
                                            except (ValueError, IndexError):
                                                pass
                        break  # 成功读取，跳出循环
                    except UnicodeDecodeError:
                        continue  # 尝试下一个编码

        print(f"从价格报告加载了 {len(price_map)} 个价格")

        # 2. 读取源文件数据（EP-0/1/2.xlsm）
        source_data = {}  # sku -> row_data

        for source_filename in source_filenames:
            print(f"正在读取源文件: {source_filename}")
            source_path = UPLOADS_DIR / source_filename
            if not source_path.exists():
                print(f"  文件不存在，跳过")
                continue

            wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
            ws = wb['Template']

            # 从第 7 行开始读取数据
            row_count = 0
            for row_idx in range(7, ws.max_row + 1):
                row = list(ws[row_idx])
                if len(row) > 2 and row[2].value:
                    sku = str(row[2].value).strip().upper()
                    # 保存整行数据（只保存有值的单元格）
                    source_data[sku] = [(cell.value if cell.value is not None else '') for cell in row]
                    row_count += 1

                    if row_count % 1000 == 0:
                        print(f"  已读取 {row_count} 行")

            wb.close()
            print(f"  完成，共读取 {row_count} 行")

        print(f"从源文件加载了 {len(source_data)} 个 SKU 数据")

        # 3. 读取模板文件
        template_path = UPLOADS_DIR / template_filename
        if not template_path.exists():
            raise FileNotFoundError(f"模板文件不存在: {template_filename}")

        wb_template = openpyxl.load_workbook(template_path)
        ws_template = wb_template['Template']

        # 4. 处理模板中的 SKU
        print(f"正在处理模板文件...")
        processed_count = 0

        # 从第 7 行开始处理（假设模板也是相同格式）
        for row_idx in range(7, ws_template.max_row + 1):
            row = list(ws_template[row_idx])
            if len(row) > 2 and row[2].value:
                sku = str(row[2].value).strip().upper()

                # 在源数据中查找
                if sku in source_data:
                    source_row = source_data[sku]

                    # 复制源数据到模板
                    for col_idx, value in enumerate(source_row):
                        if value:  # 只复制非空值
                            ws_template.cell(row=row_idx, column=col_idx + 1, value=value)

                    # 更新价格（如果有）
                    if sku in price_map:
                        # 假设价格在列 15（Your Price）
                        ws_template.cell(row=row_idx, column=16, value=price_map[sku])

                    processed_count += 1

                    if processed_count % 100 == 0:
                        print(f"  已处理 {processed_count} 行")

        print(f"处理完成，共处理 {processed_count} 行")

        # 5. 保存输出文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"processed_{timestamp}.xlsx"
        output_path = RESULTS_DIR / output_filename

        wb_template.save(output_path)
        wb_template.close()

        return output_filename, processed_count

    def process_excel(
        self,
        template_type: str,
        filenames: List[str],
        selected_prefixes: List[str],
        generated_skus: Optional[List[str]] = None,
        target_color: Optional[str] = None,
        target_size: Optional[str] = None,
        source_style_map: Optional[Dict[str, str]] = None,
        processing_mode: Optional[str] = None,
        clear_image_urls: bool = False,
        follow_sell_mode: bool = False,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> Tuple[str, int]:
        """处理 Excel 文件并生成新文件（基于 d76 逻辑，支持前端生成 SKU）"""
        def progress(message: str) -> None:
            print(message)
            if progress_callback:
                try:
                    progress_callback(message)
                except Exception:
                    pass

        job_started_at = perf_counter()
        progress(f"[DEBUG] template_type = {template_type}")
        progress(f"[DEBUG] image_variant = {TEMPLATES[template_type]['image_variant']}")

        if not filenames:
            raise ValueError("没有输入文件")

        if not selected_prefixes:
            raise ValueError("请至少输入一个产品前缀")
        normalized_prefixes = [str(p).strip().upper() for p in selected_prefixes if str(p).strip()]
        if not normalized_prefixes:
            raise ValueError("请至少输入一个有效的产品前缀")
        normalized_source_style_map = {
            str(k).strip().upper(): str(v).strip().upper()
            for k, v in (source_style_map or {}).items()
            if str(k).strip() and str(v).strip()
        }
        normalized_processing_mode = str(processing_mode or "").strip().lower()

        def prefix_matches(product_code: str) -> bool:
            # 兼容 6 位前缀（如 ES0128）和 7 位款号（如 ES0128B）
            code = str(product_code).strip().upper()
            return any(code.startswith(prefix) for prefix in normalized_prefixes)

        def parse_price_value(value: Any) -> Optional[float]:
            if value is None:
                return None
            if isinstance(value, (int, float)):
                return float(value)
            text = str(value).strip().replace("$", "").replace(",", "")
            if not text:
                return None
            try:
                return float(text)
            except (ValueError, TypeError):
                return None

        def force_price_ending_99(value: float) -> float:
            # 业务要求：Quantity Price 固定以 .99 结尾
            return round(int(value) + 0.99, 2)

        def normalize_size_display(size_value: Any) -> str:
            raw = str(size_value).strip()
            if raw.isdigit():
                return str(int(raw))
            return raw

        def replace_us_size_token(text: str, size_value: Any) -> str:
            size_display = normalize_size_display(size_value)
            return re.sub(r'\bUS\d{1,2}\b', f'US{size_display}', text, flags=re.IGNORECASE)

        def resolve_color_text(color_code: str) -> str:
            color_name = color_mapper.get_color_name(color_code)
            return color_name if color_name else color_code

        def replace_color_whole_words(text: str, source_color_code: str, target_color_code: str, *, lower_target: bool = False) -> str:
            source_text = resolve_color_text(source_color_code).strip()
            target_text = resolve_color_text(target_color_code).strip()
            if not source_text or not target_text:
                return text

            if lower_target:
                source_text = source_text.lower()
                target_text = target_text.lower()

            replaced = re.sub(
                rf'(?<![A-Za-z0-9]){re.escape(source_text)}(?![A-Za-z0-9])',
                target_text,
                text,
                flags=re.IGNORECASE,
            )

            # 兼容 mint/light/dark 等前缀颜色词，保持整词匹配
            base_word = source_text.split()[-1]
            for prefix in ("mint", "light", "dark", "bright", "deep", "pale"):
                replaced = re.sub(
                    rf'(?<![A-Za-z0-9]){prefix}\s+{re.escape(base_word)}(?![A-Za-z0-9])',
                    target_text,
                    replaced,
                    flags=re.IGNORECASE,
                )
            return replaced

        def normalize_variation_theme(value: Any) -> str:
            """Normalize variation theme to the web-facing canonical values."""
            raw = str(value or "").strip().upper()
            compact = re.sub(r"[^A-Z]", "", raw)
            if compact == "SIZECOLOR":
                return "SizeColor"
            return "SizeName-ColorName"

        field_aliases: Dict[str, List[str]] = {
            "key product features": ["Bullet Point", "Special Features", "Features"],
            "generic keyword": ["Generic Keywords"],
        }

        def read_source_value(
            source_headers: Dict[str, List[int]],
            row_values: List[Any],
            candidates: List[str],
            field_used: Optional[Dict[str, int]] = None,
        ) -> Any:
            def read_from_field(field_name: str) -> Tuple[Any, bool]:
                in_key = self._normalize_header(field_name)
                cols = source_headers.get(in_key, [])
                if not cols:
                    return None, False

                if field_used is not None:
                    in_idx = field_used[in_key]
                    if in_idx >= len(cols):
                        in_idx = len(cols) - 1
                    col_indexes = [cols[in_idx]]
                    field_used[in_key] += 1
                else:
                    col_indexes = cols

                for col_idx in col_indexes:
                    value = row_values[col_idx - 1] if 0 < col_idx <= len(row_values) else None
                    if value is not None and str(value).strip() != "":
                        return value, True
                return None, True

            for candidate in candidates:
                value, matched = read_from_field(candidate)
                if value is not None:
                    return value
                if matched:
                    continue

                for alias in field_aliases.get(self._normalize_header(candidate), []):
                    alias_value, alias_matched = read_from_field(alias)
                    if alias_value is not None:
                        progress(f"字段映射: '{candidate}' 未找到，使用别名 '{alias}' 匹配成功")
                        return alias_value
                    if alias_matched:
                        break
            return None

        def read_repeated_source_values(
            source_headers: Dict[str, List[int]],
            row_values: List[Any],
            candidates: List[str],
            count: int,
        ) -> List[Any]:
            field_used: Dict[str, int] = defaultdict(int)
            values: List[Any] = []
            for _ in range(count):
                values.append(
                    read_source_value(
                        source_headers,
                        row_values,
                        candidates,
                        field_used,
                    )
                )
            return values

        # 如果指定了目标颜色，验证颜色是否存在
        target_color_name = None
        if target_color:
            target_color = target_color.upper()
            target_color_name = color_mapper.get_color_name(target_color)
            if not target_color_name:
                raise ValueError(f"未知的目标颜色代码: {target_color}")

        # 固定输出模板（保持模板格式）
        template_file = TEMPLATES[template_type]["template_file"]
        template_path = UPLOADS_DIR / template_file
        if not template_path.exists():
            raise FileNotFoundError(f"模板文件不存在: {template_file}")
        progress(f"正在加载模板文件: {template_file}")
        template_snapshot, template_cache_hit = self._load_template_snapshot_cached(template_file)
        if template_cache_hit:
            progress("模板文件加载完成（命中内存缓存）")
        else:
            progress("模板文件加载完成（冷加载）")
        template_effective_max_col = int(template_snapshot["effective_max_col"])
        template_row_snapshots: Dict[int, Tuple[List[Any], List[Dict[str, Any]], Optional[float]]] = template_snapshot["row_snapshots"]
        template_column_widths: Dict[int, Optional[float]] = template_snapshot["column_widths"]
        progress(f"使用模板文件: {template_file}")

        # 数据源文件（忽略 txt）
        data_files = [name for name in filenames if not name.lower().endswith('.txt')]
        price_report_file = next((name for name in filenames if name.lower().endswith('.txt')), None)
        if not data_files:
            raise ValueError("缺少 Excel 数据文件")
        progress(f"数据文件: {data_files}")
        template_store = self._guess_store_from_template_type(template_type)
        sku_store_hint = self._guess_store_from_skus(generated_skus)
        preferred_store = sku_store_hint or template_store
        store_search_order = self._build_store_search_order(preferred_store)
        progress(f"店铺索引查询顺序: {' -> '.join(store_search_order)}")

        # 价格映射（All Listings Report）
        price_store = self._resolve_store_for_filename(price_report_file, preferred_store or "EP")
        price_map = self._load_price_map_cached(price_report_file, price_store)
        asin_map = self._load_asin_map_cached(price_report_file)
        asin_map_by_base: Dict[str, str] = {}
        for asin_sku, asin_value in asin_map.items():
            if len(asin_sku) >= 11 and asin_sku[:11] not in asin_map_by_base:
                asin_map_by_base[asin_sku[:11]] = asin_value

        # 创建输出工作簿
        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)
        output_ws = output_wb.create_sheet(title="Template")

        header_snapshots = {
            row_idx: template_row_snapshots[row_idx]
            for row_idx in (1, 2, 3)
        }
        prototype_snapshots = {
            row_idx: template_row_snapshots[row_idx]
            for row_idx in (4, 5)
        }

        # 复制前 3 行（表头）
        progress("正在复制表头...")
        for row_idx in range(1, 4):
            row_values, row_styles, row_height = header_snapshots[row_idx]
            for col_idx in range(1, template_effective_max_col + 1):
                target_cell = output_ws.cell(row=row_idx, column=col_idx)
                target_cell.value = row_values[col_idx - 1]
                style = row_styles[col_idx - 1]
                target_cell.font = copy(style["font"])
                target_cell.border = copy(style["border"])
                target_cell.fill = copy(style["fill"])
                target_cell.number_format = style["number_format"]
                target_cell.protection = copy(style["protection"])
                target_cell.alignment = copy(style["alignment"])
            if row_height is not None:
                output_ws.row_dimensions[row_idx].height = row_height

        # 复制列宽
        for col_idx in range(1, template_effective_max_col + 1):
            width = template_column_widths.get(col_idx)
            if width is not None:
                col_letter = get_column_letter(col_idx)
                output_ws.column_dimensions[col_letter].width = width

        progress("表头复制完成")

        # 模板表头索引（按显示名，支持重名字段）
        output_header_map: Dict[str, List[int]] = defaultdict(list)
        header_row2_values = template_row_snapshots[2][0]
        for col_idx in range(1, template_effective_max_col + 1):
            header_name = header_row2_values[col_idx - 1]
            if header_name:
                output_header_map[self._normalize_header(header_name)].append(col_idx)

        def first_output_col(*header_names: str, default: Optional[int] = None) -> int:
            for header_name in header_names:
                cols = output_header_map.get(self._normalize_header(header_name), [])
                if cols:
                    return cols[0]
            if default is None:
                raise ValueError(f"模板缺少必要列: {header_names}")
            return default

        def all_output_cols(*header_names: str) -> List[int]:
            merged: List[int] = []
            for header_name in header_names:
                merged.extend(output_header_map.get(self._normalize_header(header_name), []))
            return list(dict.fromkeys(merged))

        seller_sku_col = first_output_col("Seller SKU", default=2)
        quantity_col = first_output_col("Quantity", default=17)
        parentage_col = first_output_col("Parentage", default=83)
        product_name_col = first_output_col("Product Name", default=5)
        style_number_col = first_output_col("Style Number", default=10)
        manufacturer_part_number_col = first_output_col("Manufacturer Part Number", default=13)
        main_image_col = first_output_col("Main Image URL", default=73)
        swatch_image_col = first_output_col("Swatch Image URL", default=82)
        other_image_cols = all_output_cols("Other Image URL")
        if not other_image_cols:
            other_image_cols = list(range(74, 82))
        size_display_cols = all_output_cols("Apparel Size Value", "Size Map", "Size")
        if not size_display_cols:
            size_display_cols = [30, 299, 153]
        colour_map_col = first_output_col("Colour Map", "Color Map", default=112)
        generic_keyword_col = first_output_col("Generic Keyword", default=95)
        color_col = first_output_col("Color", default=136)
        launch_date_cols = all_output_cols("Launch Date", "product_site_launch_date")
        if not launch_date_cols:
            launch_date_cols = [532]
        release_date_cols = all_output_cols("Release Date")
        if not release_date_cols:
            release_date_cols = [520]
        quantity_price_type_cols = all_output_cols("Quantity Price Type")
        quantity_lower_bound_1_cols = all_output_cols("Quantity Lower Bound 1")
        quantity_lower_bound_2_cols = all_output_cols("Quantity Lower Bound 2")
        quantity_lower_bound_3_cols = all_output_cols("Quantity Lower Bound 3")
        item_length_description_cols = all_output_cols("Item Length Description")
        if not item_length_description_cols and template_type in {"DaMaUS", "PZUS"}:
            item_length_description_cols = [132]
        key_product_feature_cols = all_output_cols("Key Product Features", "Bullet Point")
        embellishment_feature_cols = all_output_cols("Embellishment Feature")

        # 对比映射规则（Sheet1）
        mapping_rules: List[Tuple[str, str, str]] = []
        mapping_file = UPLOADS_DIR / "对比(3).xlsx"
        if mapping_file.exists():
            progress("正在读取字段映射文件...")
            map_wb = openpyxl.load_workbook(mapping_file, data_only=True, read_only=True)
            map_ws = map_wb["Sheet1"] if "Sheet1" in map_wb.sheetnames else map_wb.active
            for r in range(2, map_ws.max_row + 1):
                out_field = map_ws.cell(row=r, column=1).value
                in_field = map_ws.cell(row=r, column=2).value
                logic = map_ws.cell(row=r, column=5).value
                if out_field:
                    mapping_rules.append((
                        str(out_field).strip(),
                        str(in_field).strip() if in_field else "",
                        str(logic).strip() if logic else "",
                    ))
            map_wb.close()
            progress(f"加载字段映射规则 {len(mapping_rules)} 条")
        else:
            progress("未找到对比(3).xlsx，使用内置最小映射")

        # 数据索引缓存
        progress("正在准备数据索引...")
        source_index_started_at = perf_counter()
        indexes_by_store = self._build_source_indexes_cached(data_files, preferred_store)
        available_store_order = [store for store in store_search_order if store in indexes_by_store]
        if not available_store_order and indexes_by_store:
            available_store_order = sorted(indexes_by_store.keys())
        if not available_store_order:
            raise ValueError("未找到可用店铺索引，请先上传 EP/DM/PZ 数据源文件")
        source_index = self._merge_source_indexes(indexes_by_store, available_store_order)
        progress(f"命中店铺索引: {', '.join(available_store_order)}")
        progress(f"数据索引准备耗时 {perf_counter() - source_index_started_at:.2f}s")
        sku_to_source: Dict[str, str] = source_index["sku_to_source"]
        sku_base_to_source: Dict[str, str] = source_index["sku_base_to_source"]
        style_suffix_to_source: Dict[Tuple[str, str], str] = source_index["style_suffix_to_source"]
        style_to_source: Dict[str, str] = source_index["style_to_source"]
        style_size_to_source: Dict[Tuple[str, str], str] = source_index["style_size_to_source"]
        style_size_suffix_to_source: Dict[Tuple[str, str, str], str] = source_index["style_size_suffix_to_source"]
        style_color_to_source: Dict[Tuple[str, str], str] = source_index["style_color_to_source"]
        style_color_suffix_to_source: Dict[Tuple[str, str, str], str] = source_index["style_color_suffix_to_source"]
        source_rows: Dict[str, List[Any]] = source_index["source_rows"]
        source_file_by_sku: Dict[str, str] = source_index["source_file_by_sku"]
        source_header_map_by_file: Dict[str, Dict[str, List[int]]] = source_index["source_header_map_by_file"]

        def source_sku_has_value(source_sku_key: str, header_names: List[str]) -> bool:
            row_values = source_rows.get(source_sku_key, [])
            source_file = source_file_by_sku.get(source_sku_key, "")
            header_map = source_header_map_by_file.get(source_file, {})
            if not row_values or not header_map:
                return False
            for header_name in header_names:
                cols = header_map.get(self._normalize_header(header_name), [])
                for col_idx in cols:
                    if 0 < col_idx <= len(row_values):
                        value = row_values[col_idx - 1]
                        if value is not None and str(value).strip() != "":
                            return True
            return False

        def prefer_with_product_id(existing_sku: Optional[str], candidate_sku: str) -> str:
            if existing_sku is None:
                return candidate_sku
            existing_has_pid = source_sku_has_value(existing_sku, ["Product ID", "External Product ID"])
            candidate_has_pid = source_sku_has_value(candidate_sku, ["Product ID", "External Product ID"])
            if candidate_has_pid and not existing_has_pid:
                return candidate_sku
            return existing_sku

        style_color_size_to_source: Dict[Tuple[str, str, str], str] = {}
        style_color_size_suffix_to_source: Dict[Tuple[str, str, str, str], str] = {}
        for source_sku in sku_to_source.keys():
            parsed_source = self.parse_sku(source_sku)
            if not parsed_source:
                continue
            color_size_key = (parsed_source.product_code, parsed_source.color_code, parsed_source.size)
            color_size_suffix_key = (parsed_source.product_code, parsed_source.color_code, parsed_source.size, parsed_source.suffix)
            style_color_size_to_source[color_size_key] = prefer_with_product_id(
                style_color_size_to_source.get(color_size_key),
                source_sku,
            )
            style_color_size_suffix_to_source[color_size_suffix_key] = prefer_with_product_id(
                style_color_size_suffix_to_source.get(color_size_suffix_key),
                source_sku,
            )

        # 待处理 SKU 列表：优先前端生成
        all_skus: List[str] = []
        if generated_skus:
            dedup = set()
            for raw_sku in generated_skus:
                sku = str(raw_sku).strip().upper()
                if len(sku) < 11 or sku in dedup:
                    continue
                info = self.parse_sku(sku)
                if info and prefix_matches(info.product_code):
                    all_skus.append(sku)
                    dedup.add(sku)
            progress(f"使用前端生成 SKU，共 {len(all_skus)} 个")
        else:
            for source_sku in sku_to_source.keys():
                info = self.parse_sku(source_sku)
                if info and prefix_matches(info.product_code):
                    all_skus.append(source_sku)
            progress(f"从源文件筛选 SKU，共 {len(all_skus)} 个")

        # 匹配模式识别：
        # 加色（一码多色）=> 前7位+后2位(size)匹配
        # 加码（一色多码）=> 前9位(style+color)匹配
        parsed_for_mode = [self.parse_sku(s) for s in all_skus]
        parsed_for_mode = [p for p in parsed_for_mode if p]
        unique_colors = {p.color_code for p in parsed_for_mode}
        unique_sizes = {p.size for p in parsed_for_mode}
        match_mode = "mixed"
        if len(unique_sizes) == 1 and len(unique_colors) > 1:
            match_mode = "add-color"
        elif len(unique_colors) == 1 and len(unique_sizes) > 1:
            match_mode = "add-code"
        requested_match_mode = (
            normalized_processing_mode
            if normalized_processing_mode in {"add-color", "add-code"}
            else ""
        )
        effective_match_mode = requested_match_mode if requested_match_mode else match_mode
        if requested_match_mode == "add-code" and len(unique_sizes) == 1 and len(unique_colors) > 1:
            effective_match_mode = "add-color"
            progress("检测到一码多色，自动切换为 add-color 匹配")
        progress(f"源匹配模式: {match_mode} (colors={len(unique_colors)}, sizes={len(unique_sizes)})")
        if requested_match_mode and effective_match_mode == requested_match_mode and effective_match_mode != match_mode:
            progress(f"按请求模式覆盖匹配策略: {effective_match_mode}")
        elif requested_match_mode and effective_match_mode != requested_match_mode:
            progress(f"请求模式 {requested_match_mode} 与 SKU 结构不一致，已改为 {effective_match_mode}")

        base_suffix, plus_suffix = self._get_store_suffix_family(template_type)

        def suffix_candidates(suffix: str) -> List[str]:
            """返回当前店铺下可接受的后缀回退顺序。"""
            normalized_suffix = (suffix or "").strip().upper()
            if normalized_suffix and not normalized_suffix.startswith("-"):
                normalized_suffix = f"-{normalized_suffix}"

            if normalized_suffix == base_suffix:
                candidates = [base_suffix, plus_suffix, ""]
            elif normalized_suffix == plus_suffix:
                candidates = [plus_suffix, base_suffix, ""]
            elif normalized_suffix:
                candidates = [normalized_suffix, base_suffix, plus_suffix, ""]
            else:
                candidates = [base_suffix, plus_suffix, ""]

            ordered: List[str] = []
            seen = set()
            for candidate in candidates:
                if candidate in seen:
                    continue
                seen.add(candidate)
                ordered.append(candidate)
            return ordered

        # 无后缀请求展开规则：
        # - 跟卖：展开为当前店铺的标准/Plus 两类后缀
        # - 加码：若同款同色在源数据里同时存在两套 listing，则双展开
        # - 其他加色/加码：仅展开为当前尺码对应的标准后缀，避免误生成其他族后缀
        expanded_skus: List[str] = []
        expanded_seen = set()
        for requested_sku in all_skus:
            req_info = self.parse_sku(requested_sku)
            if not req_info:
                continue
            if req_info.suffix:
                targets = [req_info.suffix]
            elif follow_sell_mode:
                targets = [base_suffix, plus_suffix]
            elif requested_match_mode == "add-code":
                dual_suffix_targets = [
                    suffix
                    for suffix in (base_suffix, plus_suffix)
                    if style_color_suffix_to_source.get((req_info.product_code, req_info.color_code, suffix))
                ]
                targets = dual_suffix_targets or [self._generate_suffix(template_type, target_size or req_info.size)]
            else:
                generated_suffix = self._generate_suffix(template_type, target_size or req_info.size)
                targets = [generated_suffix]
            for sfx in targets:
                candidate = f"{req_info.product_code}{req_info.color_code}{req_info.size}{sfx}"
                if candidate not in expanded_seen:
                    expanded_seen.add(candidate)
                    expanded_skus.append(candidate)

        processed_count = 0
        output_row_idx = 4
        add_color_display_source_refs: Dict[Tuple[str, str], str] = {}
        add_color_display_price_cache: Dict[str, float] = {}

        process_rows_started_at = perf_counter()
        progress(f"正在处理 {len(expanded_skus)} 个 SKU...")
        for sku in expanded_skus:
            info = self.parse_sku(sku)
            if not info:
                continue
            source_style = normalized_source_style_map.get(info.product_code, info.product_code)

            source_ref = None
            if follow_sell_mode:
                # 跟卖必须按老款一对一匹配，避免 Product ID/Color 串值
                for sfx in suffix_candidates(info.suffix):
                    source_ref = style_color_size_suffix_to_source.get((source_style, info.color_code, info.size, sfx))
                    if source_ref:
                        break
                if source_ref is None:
                    source_ref = style_color_size_to_source.get((source_style, info.color_code, info.size))
            else:
                # 优先完整 SKU 匹配，再回退到 11 位基准匹配
                source_ref = sku_to_source.get(sku)
                if source_ref is None and len(sku) >= 11:
                    source_ref = sku_base_to_source.get(sku[:11])
                allow_generic_fallback = True

                if source_ref is None and effective_match_mode == "add-color":
                    # 加色：前7位 + 后2位(size)，并兼容当前店铺后缀族
                    for sfx in suffix_candidates(info.suffix):
                        source_ref = style_size_suffix_to_source.get((source_style, info.size, sfx))
                        if source_ref:
                            break
                    if source_ref is None:
                        source_ref = style_size_to_source.get((source_style, info.size))
                elif source_ref is None and effective_match_mode == "add-code":
                    # 加码：前9位(style+color)，并兼容当前店铺后缀族
                    for sfx in suffix_candidates(info.suffix):
                        source_ref = style_color_suffix_to_source.get((source_style, info.color_code, sfx))
                        if source_ref:
                            break
                    if source_ref is None:
                        source_ref = style_color_to_source.get((source_style, info.color_code))
                    # 加码必须保持同色，禁止回退到无颜色约束匹配
                    allow_generic_fallback = False

                # 通用回退：前7位+后缀，再退化到前7位
                if source_ref is None and allow_generic_fallback:
                    source_ref = style_suffix_to_source.get((source_style, info.suffix))
                if source_ref is None and allow_generic_fallback:
                    source_ref = style_to_source.get(source_style)
            if source_ref is None:
                continue
            source_sku = source_ref
            source_row_values = source_rows.get(source_sku, [])
            source_file = source_file_by_sku.get(source_sku, "")
            source_header_map = source_header_map_by_file.get(source_file, {})
            source_info = self.parse_sku(source_sku)
            use_canonical_add_color_display = (
                not follow_sell_mode and normalized_processing_mode == "add-color"
            )
            display_source_sku = source_sku
            display_source_row_values = source_row_values
            display_source_header_map = source_header_map
            display_source_info = source_info
            if use_canonical_add_color_display:
                display_key = (
                    source_style,
                    "plus" if self._is_plus_body_suffix(info.suffix) else "base",
                )
                display_source_sku = add_color_display_source_refs.setdefault(display_key, source_sku)
                display_source_row_values = source_rows.get(display_source_sku, source_row_values)
                display_source_file = source_file_by_sku.get(display_source_sku, source_file)
                display_source_header_map = source_header_map_by_file.get(display_source_file, source_header_map)
                display_source_info = self.parse_sku(display_source_sku)

            # 按源 SKU 后缀选择模板原型行：Plus Body 用第4行，常规款用第5行
            prototype_row_idx = 4 if self._is_plus_body_suffix(source_info.suffix if source_info else "") else 5

            # 先复制模板原型行（保证输出结构完整）
            prototype_values, prototype_styles, prototype_height = prototype_snapshots[prototype_row_idx]
            for col_idx in range(1, template_effective_max_col + 1):
                target_cell = output_ws.cell(row=output_row_idx, column=col_idx)
                target_cell.value = prototype_values[col_idx - 1]
                style = prototype_styles[col_idx - 1]
                target_cell.font = copy(style["font"])
                target_cell.border = copy(style["border"])
                target_cell.fill = copy(style["fill"])
                target_cell.number_format = style["number_format"]
                target_cell.protection = copy(style["protection"])
                target_cell.alignment = copy(style["alignment"])

            if prototype_height is not None:
                output_ws.row_dimensions[output_row_idx].height = prototype_height

            # 根据对比映射逐行覆盖（支持重复字段顺序映射）
            out_field_used: Dict[str, int] = defaultdict(int)
            in_field_used: Dict[str, int] = defaultdict(int)

            if mapping_rules:
                for out_field, in_field, logic in mapping_rules:
                    out_key = self._normalize_header(out_field)
                    out_cols = output_header_map.get(out_key, [])
                    if not out_cols:
                        continue

                    out_idx = out_field_used[out_key]
                    if out_idx >= len(out_cols):
                        out_idx = len(out_cols) - 1
                    out_col = out_cols[out_idx]
                    out_field_used[out_key] += 1

                    # 图片字段特殊逻辑：不从输入表覆盖图片 URL，按模板原型/后续图片规则处理
                    if "图片链接替换" in logic:
                        continue

                    # 替换/颜色尺码变更：从输入表字段取值
                    if ("替换" in logic or "颜色/尺码变更" in logic) and in_field and in_field not in ("非映射", "-"):
                        value = read_source_value(
                            source_header_map,
                            source_row_values,
                            [in_field],
                            in_field_used,
                        )
                        if value is not None:
                            output_ws.cell(row=output_row_idx, column=out_col).value = value
                        continue

                    # 固定值/默认值
                    if "默认数字5" in logic:
                        output_ws.cell(row=output_row_idx, column=out_col).value = 5
                    elif "填写固定值“Child”" in logic or "固定“Child”" in logic:
                        output_ws.cell(row=output_row_idx, column=out_col).value = "Child"
            else:
                # 兜底最小映射
                fallback = [
                    ("product type", "product type"),
                    ("seller sku", "sku"),
                    ("brand name", "brand name"),
                    ("product name", "item name"),
                    ("product description", "product description"),
                    ("item type keyword", "item type keyword"),
                ]
                for out_key, in_key in fallback:
                    out_cols = output_header_map.get(out_key, [])
                    in_cols = source_header_map.get(in_key, [])
                    if out_cols and in_cols:
                        in_col = in_cols[0]
                        value = source_row_values[in_col - 1] if 0 < in_col <= len(source_row_values) else None
                        if value is not None and str(value).strip() != "":
                            output_ws.cell(row=output_row_idx, column=out_cols[0]).value = value

            # Product Type 统一小写（与模板样例一致）
            product_type_cols = output_header_map.get("product type", [])
            if product_type_cols:
                product_type_val = output_ws.cell(row=output_row_idx, column=product_type_cols[0]).value
                if product_type_val:
                    output_ws.cell(row=output_row_idx, column=product_type_cols[0]).value = str(product_type_val).lower()

            # 基础字段
            output_ws.cell(row=output_row_idx, column=seller_sku_col).value = sku
            zero_quantity_for_store_add_color = (
                not follow_sell_mode
                and normalized_processing_mode == "add-color"
                and template_type in {"DaMaUS", "PZUS"}
            )
            output_ws.cell(row=output_row_idx, column=quantity_col).value = (
                0 if (follow_sell_mode or zero_quantity_for_store_add_color) else 5
            )
            output_ws.cell(row=output_row_idx, column=parentage_col).value = "Child"
            # 批发阶梯固定值：所有模式统一固定
            for col in quantity_price_type_cols:
                output_ws.cell(row=output_row_idx, column=col).value = "Fixed"
            for col in quantity_lower_bound_1_cols:
                output_ws.cell(row=output_row_idx, column=col).value = 3
            for col in quantity_lower_bound_2_cols:
                output_ws.cell(row=output_row_idx, column=col).value = 5
            for col in quantity_lower_bound_3_cols:
                output_ws.cell(row=output_row_idx, column=col).value = 8

            # 图片字段：默认严格沿用模板原型行（与正确样例一致）
            image_copy_cols = [main_image_col, *other_image_cols, swatch_image_col]
            for col_idx in dict.fromkeys(image_copy_cols):
                if col_idx <= len(prototype_values):
                    output_ws.cell(row=output_row_idx, column=col_idx).value = prototype_values[col_idx - 1]
                else:
                    output_ws.cell(row=output_row_idx, column=col_idx).value = None
            output_ws.cell(row=output_row_idx, column=swatch_image_col).value = output_ws.cell(
                row=output_row_idx, column=main_image_col
            ).value

            # 尺码字段：始终按目标 SKU 同步（不依赖 target_size 模式）
            size_without_leading_zero = str(int(info.size))
            for col in size_display_cols:
                output_ws.cell(row=output_row_idx, column=col).value = size_without_leading_zero

            # 产品名中的 US 尺码与目标 SKU 对齐
            name_cell = output_ws.cell(row=output_row_idx, column=product_name_col)
            if name_cell.value:
                name_cell.value = replace_us_size_token(str(name_cell.value), info.size)

            # Outer Material Type：仅填充第一列，后续重复列保持空白
            outer_material_cols = output_header_map.get("outer material type", [])
            if outer_material_cols:
                material_candidates = [
                    [f"material{i}", f"material {i}"]
                    for i in range(1, 6)
                ]
                first_outer_material_col = outer_material_cols[0]
                first_material_candidates = material_candidates[0]
                material_value = read_source_value(
                    source_header_map,
                    source_row_values,
                    first_material_candidates,
                )
                output_ws.cell(row=output_row_idx, column=first_outer_material_col).value = (
                    material_value if material_value is not None else None
                )
                for out_col in outer_material_cols[1:]:
                    output_ws.cell(row=output_row_idx, column=out_col).value = None

            # 修复 1：统一 Material 字段权威值。
            # 所有店铺模板统一优先取源表 Fabric Type，其次回退到第一个非空的 Material1~5，
            # 再同步覆盖 Fabric Type / Outer Material Type，避免模板默认值或重复映射导致两者不一致。
            material_authority_value = read_source_value(
                source_header_map,
                source_row_values,
                [
                    "Fabric Type",
                    "Material1",
                    "Material 1",
                    "Material2",
                    "Material 2",
                    "Material3",
                    "Material 3",
                    "Material4",
                    "Material 4",
                    "Material5",
                    "Material 5",
                ],
            )
            if material_authority_value is not None:
                fabric_type_cols = output_header_map.get("fabric type", [])
                authority_target_cols = [*fabric_type_cols]
                if outer_material_cols:
                    authority_target_cols.append(outer_material_cols[0])
                for col in dict.fromkeys(authority_target_cols):
                    output_ws.cell(row=output_row_idx, column=col).value = material_authority_value

            # 修复 2：Item Length 保持空白，不影响 Item Length Description 的正常来源值。
            item_length_cols = output_header_map.get("item length", [])
            if item_length_cols:
                for col in item_length_cols:
                    output_ws.cell(row=output_row_idx, column=col).value = None

            # 覆盖 Your Price：优先价格报告，缺失时回退源表价格列
            your_price_cols = output_header_map.get("your price", [])
            if not your_price_cols:
                your_price_cols = output_header_map.get("price", [])

            if use_canonical_add_color_display:
                source_price_value = add_color_display_price_cache.get(display_source_sku)
                if source_price_value is None:
                    source_price_value = price_map.get(display_source_sku)
                if source_price_value is None:
                    source_price_raw = read_source_value(
                        display_source_header_map,
                        display_source_row_values,
                        ["Your Price", "Price", "Standard Price"],
                        None,
                    )
                    source_price_value = parse_price_value(source_price_raw)
                if source_price_value is None and your_price_cols:
                    source_price_value = parse_price_value(
                        output_ws.cell(row=output_row_idx, column=your_price_cols[0]).value
                    )
                if source_price_value is None:
                    source_price_value = price_map.get(source_sku)
                if source_price_value is None:
                    source_price_value = price_map.get(sku)
                if source_price_value is None:
                    source_price_raw = read_source_value(
                        source_header_map,
                        source_row_values,
                        ["Your Price", "Price", "Standard Price"],
                        None,
                    )
                    source_price_value = parse_price_value(source_price_raw)
                if source_price_value is not None:
                    add_color_display_price_cache.setdefault(display_source_sku, source_price_value)
            else:
                source_price_value = price_map.get(source_sku)
                if source_price_value is None:
                    source_price_value = price_map.get(sku)
                if source_price_value is None:
                    source_price_raw = read_source_value(
                        source_header_map,
                        source_row_values,
                        ["Your Price", "Price", "Standard Price"],
                        None,
                    )
                    source_price_value = parse_price_value(source_price_raw)

            if your_price_cols and source_price_value is not None:
                output_ws.cell(row=output_row_idx, column=your_price_cols[0]).value = float(source_price_value)

            # 价格计算
            your_price_col = your_price_cols[0] if your_price_cols else 16
            your_price_raw = output_ws.cell(row=output_row_idx, column=your_price_col).value
            your_price = parse_price_value(your_price_raw)
            if your_price is not None:
                list_price_cols = output_header_map.get("list price", [513])
                business_price_cols = output_header_map.get("business price", [537])
                quantity_price_cols = output_header_map.get("quantity price 1", output_header_map.get("quantity price", [540]))
                quantity_price_2_cols = output_header_map.get("quantity price 2", [542])
                quantity_price_3_cols = output_header_map.get("quantity price 3", [544])
                if follow_sell_mode:
                    # 跟卖 SOP：
                    # 6) price = 老版本价格 - 0.1
                    # 7) list price = 新价格 + 10
                    new_price = round(your_price - 0.1, 2)
                    output_ws.cell(row=output_row_idx, column=your_price_col).value = new_price
                    list_price = round(new_price + 10, 2)
                    for col in list_price_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = list_price
                    # 补齐 B2B 价格区，避免模板黄色区域为空
                    business_price = round(new_price - 1, 2)
                    quantity_price_1 = force_price_ending_99(business_price * 0.95)
                    quantity_price_2 = force_price_ending_99(business_price * 0.92)
                    quantity_price_3 = force_price_ending_99(business_price * 0.90)
                    for col in business_price_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = business_price
                    for col in quantity_price_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = quantity_price_1
                    for col in quantity_price_2_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = quantity_price_2
                    for col in quantity_price_3_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = quantity_price_3
                else:
                    # 普通加色加码价格逻辑
                    list_price = round(your_price + 10, 2)
                    business_price = round(your_price - 1, 2)
                    quantity_price_1 = force_price_ending_99(business_price * 0.95)
                    quantity_price_2 = force_price_ending_99(business_price * 0.92)
                    quantity_price_3 = force_price_ending_99(business_price * 0.90)

                    for col in list_price_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = list_price
                    for col in business_price_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = business_price
                    for col in quantity_price_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = quantity_price_1
                    for col in quantity_price_2_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = quantity_price_2
                    for col in quantity_price_3_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = quantity_price_3

            # 所有价格相关字段统一格式
            for header_key, cols in output_header_map.items():
                if "price" not in header_key:
                    continue
                for col in cols:
                    output_ws.cell(row=output_row_idx, column=col).number_format = "0.00"

            # 所有浮点数统一两位小数，避免模板残留格式导致显示一位小数
            for col_idx in range(1, template_effective_max_col + 1):
                value_cell = output_ws.cell(row=output_row_idx, column=col_idx)
                value = value_cell.value
                if isinstance(value, bool):
                    continue
                if isinstance(value, (int, float)):
                    current_format = str(value_cell.number_format or "")
                    if isinstance(value, float) or "0.00" in current_format or "0.0" in current_format:
                        value_cell.value = round(float(value), 2)
                        value_cell.number_format = "0.00"
                    continue
                if isinstance(value, str):
                    text = value.strip()
                    if "." not in text:
                        continue
                    normalized = text.replace(",", "").replace("$", "")
                    try:
                        decimal_value = float(normalized)
                    except (ValueError, TypeError):
                        continue
                    value_cell.value = round(decimal_value, 2)
                    value_cell.number_format = "0.00"

            # Launch Date: 优先按表头名定位，避免模板列位变化导致写错列
            launch_date_str = self.calculate_launch_date()
            launch_date_dt: Optional[datetime] = None
            try:
                launch_date_dt = datetime.strptime(launch_date_str, "%Y-%m-%d")
            except ValueError:
                launch_date_dt = None

            for col in dict.fromkeys(launch_date_cols + release_date_cols):
                date_cell = output_ws.cell(row=output_row_idx, column=col)
                date_cell.value = launch_date_dt if launch_date_dt else launch_date_str
                date_cell.number_format = "yyyy/m/d"

            # 统一计算最终颜色/尺码（支持同时加色+加码）
            final_color_code = target_color if target_color else info.color_code
            final_size = target_size if target_size else info.size
            final_suffix = info.suffix if info.suffix else self._generate_suffix(template_type, final_size)
            final_suffix = final_suffix.upper()
            final_sku = f"{info.product_code}{final_color_code}{final_size}{final_suffix}"
            source_parent_sku = read_source_value(
                source_header_map,
                source_row_values,
                [
                    "Parent SKU",
                    "child_parent_sku_relationship[marketplace_id=ATVPDKIKX0DER]#1.parent_sku",
                ],
            )
            parent_sku = (
                str(source_parent_sku).strip()
                if source_parent_sku is not None and str(source_parent_sku).strip()
                else f"{info.product_code}{final_suffix}"
            )

            # 图片 URL：
            # 1) add-code 直接沿用匹配到的源行图片，不重建
            # 2) add-color / 其他场景在颜色变化时按目标 SKU 重建图片链接
            # 3) 若模板原型行图片为空，也自动补图，避免 Child 行缺图
            # 4) 自动生成时按店铺规则生成 1-4 图，Other Image URL5 固定放尺码图
            source_color_code = source_info.color_code if source_info else info.color_code
            color_changed = final_color_code != source_color_code
            info_for_image = source_info if source_info else info
            current_main_image_url = output_ws.cell(row=output_row_idx, column=main_image_col).value
            use_source_images_for_add_code = (
                not follow_sell_mode
                and (
                    normalized_processing_mode == "add-code"
                    or (not normalized_processing_mode and match_mode == "add-code")
                )
            )
            if use_source_images_for_add_code and not clear_image_urls:
                source_main_image_url = read_source_value(
                    source_header_map,
                    source_row_values,
                    ["Main Image URL", "main_image_url"],
                )
                source_swatch_image_url = read_source_value(
                    source_header_map,
                    source_row_values,
                    ["Swatch Image URL", "swatch_image_url"],
                )
                source_other_image_urls: List[str] = []
                source_other_cols: List[int] = []
                seen_other_cols = set()

                for header_name in ["Other Image URL"]:
                    for col_idx in source_header_map.get(self._normalize_header(header_name), []):
                        if col_idx not in seen_other_cols:
                            seen_other_cols.add(col_idx)
                            source_other_cols.append(col_idx)

                for idx in range(1, 6):
                    numbered_headers = [
                        f"Other Image URL {idx}",
                        f"other_image_url{idx}",
                    ]
                    for header_name in numbered_headers:
                        for col_idx in source_header_map.get(self._normalize_header(header_name), []):
                            if col_idx not in seen_other_cols:
                                seen_other_cols.add(col_idx)
                                source_other_cols.append(col_idx)

                source_other_cols.sort()
                for col_idx in source_other_cols:
                    if 0 < col_idx <= len(source_row_values):
                        value = source_row_values[col_idx - 1]
                        if value is None:
                            continue
                        text = str(value).strip()
                        if text:
                            source_other_image_urls.append(text)

                output_ws.cell(row=output_row_idx, column=main_image_col).value = source_main_image_url
                for col in other_image_cols:
                    output_ws.cell(row=output_row_idx, column=col).value = None
                for idx, url in enumerate(source_other_image_urls):
                    if idx < len(other_image_cols):
                        output_ws.cell(row=output_row_idx, column=other_image_cols[idx]).value = url
                output_ws.cell(row=output_row_idx, column=swatch_image_col).value = (
                    source_swatch_image_url if source_swatch_image_url is not None else source_main_image_url
                )
            else:
                needs_image_fill = color_changed or not current_main_image_url
                if needs_image_fill and not clear_image_urls:
                    # 图片规则按“目标 Seller SKU 后缀”判定，不能用源行后缀。
                    is_ph_suffix = self._is_plus_body_suffix(final_suffix)
                    image_variant = TEMPLATES[template_type]["image_variant"]
                    size_chart_url = "https://eppic.s3.amazonaws.com/MH00000-S2.jpg"

                    if is_ph_suffix:
                        main_image_url = f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}101.jpg"
                        other_image_urls = [
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}102.jpg",
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}103.jpg",
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}104.jpg",
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}105.jpg",
                        ]
                    else:
                        main_image_url = f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}10.jpg"
                        other_image_urls = [
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}20.jpg",
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}30.jpg",
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}40.jpg",
                            f"https://eppic.s3.amazonaws.com/{info_for_image.product_code}{final_color_code}-{image_variant}50.jpg",
                        ]

                    output_ws.cell(row=output_row_idx, column=main_image_col).value = main_image_url
                    for col in other_image_cols:
                        output_ws.cell(row=output_row_idx, column=col).value = None
                    for idx, url in enumerate(other_image_urls):
                        if idx < len(other_image_cols):
                            output_ws.cell(row=output_row_idx, column=other_image_cols[idx]).value = url
                    # Other Image URL5 放尺码图
                    if len(other_image_cols) >= 5:
                        output_ws.cell(row=output_row_idx, column=other_image_cols[4]).value = size_chart_url
                output_ws.cell(row=output_row_idx, column=swatch_image_col).value = output_ws.cell(
                    row=output_row_idx, column=main_image_col
                ).value

            if clear_image_urls:
                # 跟卖导出要求：清空 main_image_url 与 other_image_url1~5
                for col_idx in [main_image_col, *other_image_cols[:5]]:
                    output_ws.cell(row=output_row_idx, column=col_idx).value = None
                output_ws.cell(row=output_row_idx, column=swatch_image_col).value = None

            source_variation_theme = read_source_value(
                source_header_map,
                source_row_values,
                ["Variation Theme", "Variation Theme Name"],
            )
            force_size_name_color_theme = (
                not follow_sell_mode
                and normalized_processing_mode == "add-color"
                and template_type in {"DaMaUS", "PZUS"}
            )
            normalized_variation_theme = (
                "SizeName-ColorName"
                if force_size_name_color_theme
                else normalize_variation_theme(source_variation_theme)
            )

            # Color / Colour Map：
            # SizeColor 变体保留源值，其他变体按最终颜色码重算。
            final_color_name = color_mapper.get_color_name(final_color_code)
            preserve_source_color = (
                normalized_variation_theme == "SizeColor"
                and not force_size_name_color_theme
            )

            source_color_value = None
            source_colour_map_value = None
            if preserve_source_color:
                source_color_value = read_source_value(
                    source_header_map,
                    source_row_values,
                    ["color"],
                )
                source_colour_map_value = read_source_value(
                    source_header_map,
                    source_row_values,
                    ["colour map", "color map"],
                )

            if preserve_source_color and source_color_value is not None:
                output_ws.cell(row=output_row_idx, column=color_col).value = source_color_value
            elif final_color_name:
                output_ws.cell(row=output_row_idx, column=color_col).value = final_color_name

            if preserve_source_color and source_colour_map_value is not None:
                output_ws.cell(row=output_row_idx, column=colour_map_col).value = source_colour_map_value
            elif final_color_name:
                colour_map = self.get_color_map_value(final_color_name)
                output_ws.cell(row=output_row_idx, column=colour_map_col).value = colour_map

            # Generic Keyword
            generic_keyword_cell = output_ws.cell(row=output_row_idx, column=generic_keyword_col)
            source_generic_keyword = read_source_value(
                source_header_map,
                source_row_values,
                ["generic keyword"],
            )
            if source_generic_keyword is not None:
                generic_keyword_cell.value = source_generic_keyword
            if not follow_sell_mode and generic_keyword_cell.value:
                generic_keyword = str(generic_keyword_cell.value)
                if color_changed:
                    generic_keyword = replace_color_whole_words(
                        generic_keyword,
                        source_color_code,
                        final_color_code,
                        lower_target=True,
                    )
                output_ws.cell(row=output_row_idx, column=generic_keyword_col).value = generic_keyword

            # 最终 SKU（防止颜色/尺码分步覆盖）
            output_ws.cell(row=output_row_idx, column=seller_sku_col).value = final_sku
            output_ws.cell(row=output_row_idx, column=style_number_col).value = final_sku
            output_ws.cell(row=output_row_idx, column=manufacturer_part_number_col).value = final_sku
            parent_sku_cols = output_header_map.get("parent sku", [])
            for col in parent_sku_cols:
                output_ws.cell(row=output_row_idx, column=col).value = parent_sku

            # Variation Theme：
            # SIZE/COLOR -> SizeColor，其余统一为 SizeName-ColorName。
            variation_theme_cols = output_header_map.get("variation theme", [])
            for col in variation_theme_cols:
                cell = output_ws.cell(row=output_row_idx, column=col)
                cell.value = normalized_variation_theme

            # 产品名：颜色与尺码同步到最终值
            name_cell = output_ws.cell(row=output_row_idx, column=product_name_col)
            source_product_name = read_source_value(
                display_source_header_map,
                display_source_row_values,
                ["Product Name", "Item Name"],
            )
            if source_product_name is not None:
                product_name = str(source_product_name)
            elif name_cell.value:
                product_name = str(name_cell.value)
            else:
                product_name = ""
            if product_name:
                display_source_color_code = (
                    display_source_info.color_code
                    if display_source_info
                    else (source_info.color_code if source_info else info.color_code)
                )
                if final_color_code != display_source_color_code:
                    product_name = replace_color_whole_words(
                        product_name,
                        display_source_color_code,
                        final_color_code,
                    )
                product_name = replace_us_size_token(product_name, final_size)
                name_cell.value = product_name

            source_item_length_description = read_source_value(
                display_source_header_map,
                display_source_row_values,
                ["Item Length Description"],
            )
            if source_item_length_description is not None:
                for col in item_length_description_cols:
                    output_ws.cell(row=output_row_idx, column=col).value = source_item_length_description

            if template_type == "PZUS" and use_canonical_add_color_display:
                if key_product_feature_cols:
                    feature_values = read_repeated_source_values(
                        display_source_header_map,
                        display_source_row_values,
                        ["Bullet Point", "Key Product Features"],
                        len(key_product_feature_cols),
                    )
                    for col, value in zip(key_product_feature_cols, feature_values):
                        output_ws.cell(row=output_row_idx, column=col).value = value
                if embellishment_feature_cols:
                    embellishment_values = read_repeated_source_values(
                        display_source_header_map,
                        display_source_row_values,
                        ["Embellishment Feature", "embellishment_feature1"],
                        len(embellishment_feature_cols),
                    )
                    for col, value in zip(embellishment_feature_cols, embellishment_values):
                        output_ws.cell(row=output_row_idx, column=col).value = value

            # 尺码列放在本行最后强制同步，避免被中间映射覆盖
            size_without_leading_zero = normalize_size_display(final_size)
            for col in size_display_cols:
                output_ws.cell(row=output_row_idx, column=col).value = size_without_leading_zero

            # Item Condition 固定为 New（兼容不同模板表头命名）
            item_condition_cols = []
            for header_key, cols in output_header_map.items():
                normalized_key = (header_key or "").lower()
                if ("item" in normalized_key and "condit" in normalized_key) or normalized_key == "condition type":
                    item_condition_cols.extend(cols)
            for col in dict.fromkeys(item_condition_cols):
                output_ws.cell(row=output_row_idx, column=col).value = "New"

            product_id_cols = output_header_map.get("product id", [])
            product_id_type_cols = output_header_map.get("product id type", [])
            if follow_sell_mode:
                source_product_id = (
                    asin_map.get(source_sku)
                    or asin_map.get(sku)
                    or (asin_map_by_base.get(source_sku[:11]) if len(source_sku) >= 11 else None)
                    or (
                        asin_map_by_base.get(f"{source_style}{info.color_code}{info.size}")
                        if len(source_style) == 7
                        else None
                    )
                )
                if source_product_id is None:
                    source_product_id = read_source_value(
                        source_header_map,
                        source_row_values,
                        ["Product ID", "External Product ID"],
                    )
                product_id_type_value = "ASIN"
                for col in product_id_cols:
                    output_ws.cell(row=output_row_idx, column=col).value = source_product_id
                for col in product_id_type_cols:
                    output_ws.cell(row=output_row_idx, column=col).value = product_id_type_value
            else:
                # 加色/加码模式按业务要求清空 Product ID 与 Product ID Type
                for col in product_id_cols:
                    output_ws.cell(row=output_row_idx, column=col).value = None
                for col in product_id_type_cols:
                    output_ws.cell(row=output_row_idx, column=col).value = None

            processed_count += 1
            output_row_idx += 1
            if processed_count % 100 == 0:
                progress(f"  已处理 {processed_count} 行")

        progress(f"数据处理完成，共处理 {processed_count} 行，耗时 {perf_counter() - process_rows_started_at:.2f}s")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"processed_{timestamp}.xlsx"
        output_path = RESULTS_DIR / output_filename

        progress("正在保存输出文件...")
        save_started_at = perf_counter()
        output_wb.save(output_path)
        progress(f"输出文件保存耗时 {perf_counter() - save_started_at:.2f}s")
        output_wb.close()
        progress(f"本次处理总耗时 {perf_counter() - job_started_at:.2f}s")

        return output_filename, processed_count


# 全局单例
excel_processor = ExcelProcessor()
