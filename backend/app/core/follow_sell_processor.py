"""
跟卖模块处理器
处理 SKC 输入，通过新老款映射查找尺码信息
"""
import sqlite3
import openpyxl
import re
from typing import Dict, List, Optional, Tuple, Any
from app.config import UPLOADS_DIR


class FollowSellProcessor:
    """跟卖处理器"""

    def __init__(self):
        self.new_to_old_mapping: Dict[str, str] = {}
        self.ep_data_cache: Dict[str, List[Dict]] = {}  # old_style -> [sku_data]
        self.ep_loaded = False
        self._last_parse_error = ""
        self.index_db_path = UPLOADS_DIR / "ep_index.db"
        self._load_mapping()
        # 不在初始化时加载 EP 数据，改为懒加载

    def _build_files_signature(self, filenames: List[str]) -> str:
        parts: List[str] = []
        for filename in sorted(filenames):
            path = UPLOADS_DIR / filename
            if not path.exists():
                continue
            stat = path.stat()
            parts.append(f"{filename}:{stat.st_mtime_ns}:{stat.st_size}")
        return "|".join(parts)

    def _connect_db(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.index_db_path)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        return conn

    def _ensure_index_tables(self, conn: sqlite3.Connection) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS ep_sku_index (
                style TEXT NOT NULL,
                sku TEXT NOT NULL,
                size TEXT NOT NULL,
                suffix TEXT NOT NULL,
                source_file TEXT NOT NULL,
                PRIMARY KEY (sku, source_file)
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ep_style ON ep_sku_index(style)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ep_style_size ON ep_sku_index(style, size)")
        conn.commit()

    def _ensure_ep_index(self) -> None:
        """确保 EP 索引可用，文件变化时自动重建"""
        ep_files = ['EP-0.xlsm', 'EP-1.xlsm', 'EP-2.xlsm']
        sku_pattern = re.compile(
            r"^(?P<style>[A-Z0-9]{7})(?P<color>[A-Z]{2})(?P<size>\d{2})(?P<suffix>-?[A-Z0-9]+)?$"
        )
        signature = self._build_files_signature(ep_files)
        if not signature:
            return

        conn = self._connect_db()
        try:
            self._ensure_index_tables(conn)
            row = conn.execute("SELECT value FROM meta WHERE key='ep_signature'").fetchone()
            current = row[0] if row else ""
            if current == signature:
                return

            print("EP 文件有更新，正在重建 SQLite 索引...")
            conn.execute("DELETE FROM ep_sku_index")

            insert_sql = """
                INSERT OR REPLACE INTO ep_sku_index(style, sku, size, suffix, source_file)
                VALUES (?, ?, ?, ?, ?)
            """
            total_filtered = 0

            for ep_file in ep_files:
                file_path = UPLOADS_DIR / ep_file
                if not file_path.exists():
                    print(f"警告: EP 文件不存在: {ep_file}")
                    continue
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb['Template']
                    row_count = 0
                    filtered_count = 0
                    batch: List[Tuple[str, str, str, str, str]] = []
                    for row in ws.iter_rows(min_row=7, min_col=3, max_col=3, values_only=True):
                        sku_cell = row[0] if row else None
                        if not sku_cell:
                            continue
                        sku = str(sku_cell).strip().upper()
                        matched = sku_pattern.fullmatch(sku)
                        if not matched:
                            filtered_count += 1
                            continue
                        style = matched.group("style")
                        size = matched.group("size")
                        suffix = matched.group("suffix") or ""
                        batch.append((style, sku, size, suffix, ep_file))
                        row_count += 1
                        if len(batch) >= 5000:
                            conn.executemany(insert_sql, batch)
                            conn.commit()
                            batch.clear()
                    if batch:
                        conn.executemany(insert_sql, batch)
                        conn.commit()
                    wb.close()
                    total_filtered += filtered_count
                    print(f"  {ep_file} 索引完成，共 {row_count} 行，过滤无效 SKU {filtered_count} 行")
                except Exception as e:
                    print(f"加载 {ep_file} 失败: {e}")
            print(f"EP 索引重建完成，累计过滤无效 SKU {total_filtered} 行")

            conn.execute(
                "INSERT OR REPLACE INTO meta(key, value) VALUES('ep_signature', ?)",
                (signature,)
            )
            conn.commit()
        finally:
            conn.close()

    def _load_mapping(self):
        """加载新老款映射表"""
        mapping_file = UPLOADS_DIR / "新老款映射信息(1).xlsx"

        if not mapping_file.exists():
            print(f"警告: 新老款映射文件不存在: {mapping_file}")
            return

        try:
            wb = openpyxl.load_workbook(mapping_file, read_only=True, data_only=True)
            ws = wb.active

            # 从第2行开始读取（第1行是表头）
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] and row[1]:
                    old_style = str(row[0]).strip()
                    new_style = str(row[1]).strip()
                    self.new_to_old_mapping[new_style] = old_style

            wb.close()
            print(f"加载新老款映射: {len(self.new_to_old_mapping)} 条")

        except Exception as e:
            print(f"加载新老款映射失败: {e}")

    def _load_ep_data(self):
        """兼容旧接口：现在改为确保 SQLite 索引"""
        self._ensure_ep_index()

    def parse_skc(self, skc: str) -> Optional[Tuple[str, str]]:
        """解析 SKC 格式

        Args:
            skc: SKC 字符串，格式为 7位style + 2位color，如 ES0128BDG

        Returns:
            (style, color_code) 或 None
        """
        skc = str(skc or "").strip().upper()
        self._last_parse_error = ""
        if not re.fullmatch(r"^[A-Z0-9]{7}[A-Z]{2}$", skc):
            self._last_parse_error = (
                f"SKC 格式错误: {skc or '<empty>'}，应匹配 ^[A-Z0-9]{{7}}[A-Z]{{2}}$（7位款号+2位大写字母颜色）"
            )
            return None

        style = skc[:7]  # 前7位
        color_code = skc[7:9]  # 后2位

        return style, color_code

    def _normalize_suffix(self, suffix: str) -> str:
        """按 SOP 统一 SKU 后缀。"""
        normalized = str(suffix or "").strip().upper()
        if not normalized:
            return "-USA"
        if not normalized.startswith("-"):
            normalized = f"-{normalized}"
        if normalized == "-":
            return "-USA"

        if normalized in {"-A", "-B", "-C", "-D"}:
            return "-USA"
        if normalized in {"-US", "-USB", "-USC", "-USD", "-USA"}:
            return "-USA"
        if normalized in {"-PH", "-PHB", "-PHC"}:
            return "-PH"
        return normalized

    def find_sizes_for_skc(self, skc: str) -> Dict:
        """根据 SKC 查找所有尺码信息

        Args:
            skc: SKC 字符串，如 ES0128BDG

        Returns:
            {
                'success': bool,
                'skc': str,
                'new_style': str,
                'old_style': str,
                'color_code': str,
                'sizes': [{'size': str, 'sku': str, 'suffix': str}],
                'message': str
            }
        """
        # 懒加载 EP 数据
        if not self.ep_loaded:
            self._load_ep_data()
            self.ep_loaded = True

        result = {
            'success': False,
            'skc': skc,
            'new_style': '',
            'old_style': '',
            'color_code': '',
            'sizes': [],
            'message': ''
        }

        # 1. 解析 SKC
        parsed = self.parse_skc(skc)
        if not parsed:
            result['message'] = self._last_parse_error or "SKC 格式错误"
            return result

        new_style, color_code = parsed
        result['new_style'] = new_style
        result['color_code'] = color_code

        # 2. 查找老款号
        old_style = self.new_to_old_mapping.get(new_style)

        if not old_style:
            result['message'] = f"未找到新款号 {new_style} 对应的老款号"
            return result

        result['old_style'] = old_style

        # 3. 在 EP SQLite 索引中查找老款号“同颜色”的所有尺码
        conn = self._connect_db()
        try:
            rows = conn.execute(
                """
                SELECT DISTINCT size, suffix
                FROM ep_sku_index
                WHERE style = ?
                  AND substr(sku, 8, 2) = ?
                ORDER BY size, suffix
                """,
                (old_style, color_code)
            ).fetchall()
        finally:
            conn.close()

        if not rows:
            result['message'] = f"在 EP-0/1/2 聚合表中未找到老款号 {old_style} 颜色 {color_code} 的数据"
            return result

        # 5. 构建结果
        normalized_items: List[Dict[str, str]] = []
        seen = set()
        for size, suffix in rows:
            size_str = str(size)
            suffix_str = self._normalize_suffix(str(suffix or ""))
            key = (size_str, suffix_str)
            if key in seen:
                continue
            seen.add(key)
            normalized_items.append({
                'size': size_str,
                'suffix': suffix_str,
                'sku': f"{new_style}{color_code}{size_str}{suffix_str}"  # 使用新款号构建 SKU
            })

        result['sizes'] = normalized_items

        result['success'] = True
        result['message'] = f"找到 {len(result['sizes'])} 个尺码"

        return result

    def get_sku_data_from_ep(self, old_style: str, color_code: str, size: str, suffix: str) -> Optional[List]:
        """从 EP 数据中获取指定 SKU 的完整行数据

        Args:
            old_style: 老款号（7位）
            color_code: 颜色代码（2位）
            size: 尺码（2位）
            suffix: 后缀

        Returns:
            完整行数据列表，或 None
        """
        # 构建老款号的 SKU
        target_sku = f"{old_style}{color_code}{size}{suffix}"
        self._ensure_ep_index()
        conn = self._connect_db()
        try:
            row = conn.execute(
                "SELECT source_file FROM ep_sku_index WHERE style = ? AND sku = ? LIMIT 1",
                (old_style, target_sku)
            ).fetchone()
        finally:
            conn.close()
        if not row:
            return None
        return self._load_row_data_by_sku(str(row[0]), target_sku)

    def _load_row_data_by_sku(self, ep_file: str, target_sku: str) -> Optional[List[Any]]:
        """按需从 EP 文件加载指定 SKU 的完整行数据"""
        file_path = UPLOADS_DIR / ep_file
        if not file_path.exists():
            return None

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb['Template']
            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) > 2 and row[2]:
                    sku = str(row[2]).strip()
                    if sku == target_sku:
                        wb.close()
                        return list(row)
            wb.close()
        except Exception as e:
            print(f"按需加载 SKU 行失败 {target_sku}: {e}")

        return None


# 全局单例
follow_sell_processor = FollowSellProcessor()
