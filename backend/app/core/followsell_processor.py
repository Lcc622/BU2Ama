"""
跟卖上新处理器
处理 Amazon 跟卖上新的 Excel 文件生成
"""

import logging
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class FollowSellProcessor:
    """跟卖上新处理器"""

    # 字段列映射（基于示例文件分析）
    FIELD_COLUMNS = {
        'seller_sku': 2,              # Seller SKU
        'product_id': 6,              # Product ID (ASIN)
        'product_id_type': 7,         # Product ID Type
        'style_number': 10,           # Style Number
        'manufacturer_part_number': 13,  # Manufacturer Part Number
        'your_price': 16,             # Your Price
        'quantity': 17,               # Quantity
        'release_date': 521,          # Release Date
        'launch_date': 533,           # Launch Date
    }

    def __init__(self):
        """初始化处理器"""
        self.beijing_tz = timezone(timedelta(hours=8))

    def extract_product_code(self, sku: str) -> str:
        """
        从 SKU 中提取产品代码（前7位）

        Args:
            sku: SKU 字符串，如 "ES01840BD04-PH"

        Returns:
            产品代码，如 "ES01840"

        Raises:
            ValueError: SKU 格式不正确
        """
        if not sku:
            raise ValueError("SKU 不能为空")

        # 去掉后缀（如果有）
        main_part = sku.split('-')[0] if '-' in sku else sku

        # 提取前7位作为产品代码
        if len(main_part) >= 7:
            return main_part[:7]

        raise ValueError(f"无效的 SKU 格式: {sku}")

    def calculate_launch_date(self) -> datetime:
        """
        计算上新日期（3PM 规则）

        规则：
        - 北京时间 15:00 之前：使用前一天
        - 北京时间 15:00 之后：使用当天

        Returns:
            计算后的日期（时间部分为 00:00:00）
        """
        now = datetime.now(self.beijing_tz)

        if now.hour < 15:
            # 15:00 之前，使用前一天
            date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            # 15:00 之后，使用当天
            date = now.replace(hour=0, minute=0, second=0, microsecond=0)

        logger.info(f"计算上新日期: 当前时间={now}, 使用日期={date}")
        return date

    def find_image_columns(self, header_row: tuple) -> list:
        """
        查找所有图片相关的列

        Args:
            header_row: 表头行

        Returns:
            图片列的列号列表
        """
        image_columns = []
        for idx, cell_value in enumerate(header_row, start=1):
            if cell_value and isinstance(cell_value, str):
                cell_lower = cell_value.lower()
                if 'image' in cell_lower or 'picture' in cell_lower:
                    image_columns.append(idx)

        logger.info(f"找到 {len(image_columns)} 个图片列: {image_columns}")
        return image_columns

    def find_list_price_column(self, header_row: tuple) -> int | None:
        """
        查找 list price 列

        Args:
            header_row: 表头行

        Returns:
            list price 列号，如果找不到返回 None
        """
        for idx, cell_value in enumerate(header_row, start=1):
            if cell_value and isinstance(cell_value, str):
                cell_lower = cell_value.lower()
                if 'list' in cell_lower and 'price' in cell_lower:
                    logger.info(f"找到 list price 列: {idx}")
                    return idx

        logger.warning("未找到 list price 列")
        return None

    def process(self, old_file_path: str, new_product_code: str) -> Dict[str, Any]:
        """
        处理跟卖上新

        Args:
            old_file_path: 老版本 Excel 文件路径
            new_product_code: 新产品代码，如 "ES01846"

        Returns:
            处理结果字典，包含：
            - total_skus: 处理的 SKU 总数
            - old_product_code: 识别的老产品代码
            - new_product_code: 新产品代码
            - output_filename: 生成的文件名
            - price_adjustment: 价格调整
            - date_used: 使用的日期

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式错误或产品代码无效
        """
        logger.info(f"开始处理跟卖上新: old_file={old_file_path}, new_code={new_product_code}")

        # 验证文件存在
        old_file = Path(old_file_path)
        if not old_file.exists():
            raise FileNotFoundError(f"文件不存在: {old_file_path}")

        # 验证新产品代码格式
        if not new_product_code or len(new_product_code) < 7:
            raise ValueError(f"无效的产品代码: {new_product_code}")

        # 读取老版本 Excel
        try:
            wb = openpyxl.load_workbook(old_file_path)
            sheet = wb.active
        except Exception as e:
            raise ValueError(f"无法读取 Excel 文件: {e}")

        # 获取表头（第2行）
        header_row = tuple(cell.value for cell in sheet[2])

        # 查找图片列和 list price 列
        image_columns = self.find_image_columns(header_row)
        list_price_column = self.find_list_price_column(header_row)

        # 从第一个数据行（第4行）提取老产品代码
        first_sku = sheet.cell(4, self.FIELD_COLUMNS['seller_sku']).value
        if not first_sku:
            raise ValueError("无法从第一行数据中读取 SKU")

        old_product_code = self.extract_product_code(first_sku)
        logger.info(f"识别老产品代码: {old_product_code}")

        # 计算上新日期
        launch_date = self.calculate_launch_date()

        # 统计处理的 SKU 数量
        total_skus = 0

        # 遍历所有数据行（从第4行开始）
        for row_idx in range(4, sheet.max_row + 1):
            # 读取 Seller SKU
            old_sku = sheet.cell(row_idx, self.FIELD_COLUMNS['seller_sku']).value

            # 如果 SKU 为空，跳过
            if not old_sku:
                continue

            # 替换产品代码
            new_sku = old_sku.replace(old_product_code, new_product_code)

            # 更新字段
            # 1. Seller SKU
            sheet.cell(row_idx, self.FIELD_COLUMNS['seller_sku']).value = new_sku

            # 2. Product ID (ASIN) - 保持不变
            # 不需要修改

            # 3. Product ID Type - 固定为 ASIN
            sheet.cell(row_idx, self.FIELD_COLUMNS['product_id_type']).value = "ASIN"

            # 4. Style Number
            old_style = sheet.cell(row_idx, self.FIELD_COLUMNS['style_number']).value
            if old_style:
                new_style = old_style.replace(old_product_code, new_product_code)
                sheet.cell(row_idx, self.FIELD_COLUMNS['style_number']).value = new_style

            # 5. Manufacturer Part Number
            old_part = sheet.cell(row_idx, self.FIELD_COLUMNS['manufacturer_part_number']).value
            if old_part:
                new_part = old_part.replace(old_product_code, new_product_code)
                sheet.cell(row_idx, self.FIELD_COLUMNS['manufacturer_part_number']).value = new_part

            # 6. Your Price - 减 0.1
            old_price = sheet.cell(row_idx, self.FIELD_COLUMNS['your_price']).value
            if old_price and isinstance(old_price, (int, float)):
                new_price = round(old_price - 0.1, 2)
                sheet.cell(row_idx, self.FIELD_COLUMNS['your_price']).value = new_price

                # 7. List Price - new_price + 10
                if list_price_column:
                    list_price = round(new_price + 10, 2)
                    sheet.cell(row_idx, list_price_column).value = list_price

            # 8. Quantity - 固定为 0
            sheet.cell(row_idx, self.FIELD_COLUMNS['quantity']).value = 0

            # 9. Image - 清空所有图片列
            for img_col in image_columns:
                sheet.cell(row_idx, img_col).value = None

            # 10. Release Date
            sheet.cell(row_idx, self.FIELD_COLUMNS['release_date']).value = launch_date

            # 11. Launch Date
            sheet.cell(row_idx, self.FIELD_COLUMNS['launch_date']).value = launch_date

            total_skus += 1

            if total_skus % 10 == 0:
                logger.info(f"已处理 {total_skus} 个 SKU")

        # 生成输出文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{old_file.stem}-跟卖-{new_product_code}-{timestamp}.xlsx"
        output_path = old_file.parent / output_filename

        # 保存新文件
        wb.save(output_path)
        logger.info(f"生成新文件: {output_path}")

        # 返回处理结果
        result = {
            'total_skus': total_skus,
            'old_product_code': old_product_code,
            'new_product_code': new_product_code,
            'output_filename': output_filename,
            'output_path': str(output_path),
            'price_adjustment': -0.1,
            'date_used': launch_date.strftime("%Y-%m-%d"),
        }

        logger.info(f"处理完成: {result}")
        return result
