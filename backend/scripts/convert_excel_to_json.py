#!/usr/bin/env python3
"""
将 Excel 文件的 Template sheet 转换为 JSON 格式以提高读取速度
"""
import json
import openpyxl
from pathlib import Path

def convert_excel_to_json(excel_file: Path, output_file: Path):
    """将 Excel 文件的 Template sheet 转换为 JSON"""
    print(f"正在读取: {excel_file.name}")

    wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)

    if 'Template' not in wb.sheetnames:
        print(f"  错误: 找不到 Template sheet")
        wb.close()
        return

    ws = wb['Template']

    data = []
    row_count = 0

    # 从第 7 行开始读取数据
    for row_idx in range(7, ws.max_row + 1):
        row = list(ws[row_idx])
        if len(row) > 2 and row[2].value:
            sku = str(row[2].value).strip()
            # 保存整行数据（只保存非空值）
            row_data = []
            for cell in row:
                value = cell.value
                # 转换为可 JSON 序列化的类型
                if value is None:
                    row_data.append(None)
                elif isinstance(value, (int, float, str, bool)):
                    row_data.append(value)
                else:
                    row_data.append(str(value))

            data.append({
                'sku': sku,
                'row': row_data
            })
            row_count += 1

            if row_count % 1000 == 0:
                print(f"  已处理 {row_count} 行")

    wb.close()

    print(f"正在保存到: {output_file.name}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    file_size_mb = output_file.stat().st_size / 1024 / 1024
    print(f"完成！共转换 {row_count} 行，文件大小: {file_size_mb:.2f} MB")

if __name__ == '__main__':
    # 转换 EP-0/1/2.xlsm
    uploads_dir = Path(__file__).parent.parent / 'uploads'

    for filename in ['EP-0.xlsm', 'EP-1.xlsm', 'EP-2.xlsm']:
        excel_file = uploads_dir / filename
        if excel_file.exists():
            json_file = uploads_dir / filename.replace('.xlsm', '.json')
            convert_excel_to_json(excel_file, json_file)
            print()
        else:
            print(f"文件不存在: {filename}\n")
