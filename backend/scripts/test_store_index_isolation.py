#!/usr/bin/env python3
"""
分店铺索引功能验证脚本

测试场景:
- ES01819 + NT + 04-28(偶数码)
- 店铺模板: EPUS / DaMaUS / PZUS

输出:
1) 每个店铺的结果表格（SKU / Brand Name / Product Name 前80字符 / 数据来源文件）
2) 索引库访问校验（是否只访问对应 excel_index_{store}.db）
3) 异常结论（跨店铺串值 / 数据不一致）
"""

from __future__ import annotations

import sys
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

# 允许直接从 backend/scripts 运行
BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.config import RESULTS_DIR, STORE_CONFIGS, UPLOADS_DIR  # noqa: E402
from app.core.excel_processor import excel_processor  # noqa: E402

PREFIXES = ["ES01819"]
COLORS = ["NT"]
SIZES = ["04", "06", "08", "10", "12", "14", "16", "18", "20", "22", "24", "26", "28"]
TEMPLATE_TYPES = ["EPUS", "DaMaUS", "PZUS"]

TEMPLATE_TO_STORE = {
    "EPUS": "EP",
    "DaMaUS": "DM",
    "PZUS": "PZ",
}


@dataclass
class OutputRow:
    sku: str
    brand_name: str
    product_name: str
    source_file: str = ""
    source_sku: str = ""
    expected_brand_name: str = ""
    expected_product_name: str = ""


@dataclass
class StoreRunResult:
    template_type: str
    expected_store: str
    filenames: List[str]
    output_filename: str
    processed_count: int
    db_accesses: List[str]
    rows: List[OutputRow]
    anomalies: List[str]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def product_preview(value: str, limit: int = 80) -> str:
    text = normalize_text(value)
    return text[:limit]


def build_generated_skus(prefixes: List[str], colors: List[str], sizes: List[str]) -> List[str]:
    skus: List[str] = []
    for prefix in prefixes:
        for color in colors:
            for size in sizes:
                skus.append(f"{prefix.upper()}{color.upper()}{size}")
    return skus


def resolve_filenames(template_type: str) -> List[str]:
    config = STORE_CONFIGS.get(template_type, STORE_CONFIGS["EPUS"])
    result: List[str] = []
    for source_file in config.get("source_files", []):
        if (UPLOADS_DIR / source_file).exists():
            result.append(source_file)
    listing_report = str(config.get("listing_report", "")).strip()
    if listing_report and (UPLOADS_DIR / listing_report).exists():
        result.append(listing_report)
    return result


def find_header_col(header_row: List[Any], candidates: List[str]) -> Optional[int]:
    normalized_candidates = {excel_processor._normalize_header(item) for item in candidates}
    for idx, value in enumerate(header_row, start=1):
        if value is None:
            continue
        if excel_processor._normalize_header(value) in normalized_candidates:
            return idx
    return None


def read_output_rows(output_path: Path) -> List[OutputRow]:
    wb = openpyxl.load_workbook(output_path, data_only=True)
    try:
        ws = wb["Template"] if "Template" in wb.sheetnames else wb.active
        header_row_values = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
        sku_col = find_header_col(header_row_values, ["Seller SKU", "SKU"])
        brand_col = find_header_col(header_row_values, ["Brand Name", "Brand"])
        product_col = find_header_col(header_row_values, ["Product Name", "Item Name"])

        if not sku_col or not brand_col or not product_col:
            raise RuntimeError("输出文件缺少必要列：Seller SKU / Brand Name / Product Name")

        rows: List[OutputRow] = []
        for row_idx in range(4, ws.max_row + 1):
            sku = normalize_text(ws.cell(row=row_idx, column=sku_col).value).upper()
            if not sku:
                continue
            rows.append(
                OutputRow(
                    sku=sku,
                    brand_name=normalize_text(ws.cell(row=row_idx, column=brand_col).value),
                    product_name=normalize_text(ws.cell(row=row_idx, column=product_col).value),
                )
            )
        return rows
    finally:
        wb.close()


def suffix_candidates(suffix: str) -> List[str]:
    normalized = normalize_text(suffix).upper()
    if normalized == "-USA":
        return ["-USA", "-PH", ""]
    if normalized == "-PH":
        return ["-PH", "-USA", ""]
    return ["-USA", "-PH", ""]


def normalize_size_display(size_value: str) -> str:
    raw = normalize_text(size_value)
    if raw.isdigit():
        return str(int(raw))
    return raw


def align_product_name_size(text: str, sku: str) -> str:
    info = excel_processor.parse_sku(sku)
    if not info:
        return text
    size_display = normalize_size_display(info.size)
    return re.sub(r"\bUS\d{1,2}\b", f"US{size_display}", text, flags=re.IGNORECASE)


def resolve_source_for_add_code(
    sku: str,
    source_index: Dict[str, Any],
) -> Optional[str]:
    info = excel_processor.parse_sku(sku)
    if not info:
        return None

    source_style = info.product_code
    sku_to_source: Dict[str, str] = source_index["sku_to_source"]
    sku_base_to_source: Dict[str, str] = source_index["sku_base_to_source"]
    style_color_suffix_to_source: Dict[Tuple[str, str, str], str] = source_index["style_color_suffix_to_source"]
    style_color_to_source: Dict[Tuple[str, str], str] = source_index["style_color_to_source"]

    source_ref = sku_to_source.get(sku)
    if source_ref is None and len(sku) >= 11:
        source_ref = sku_base_to_source.get(sku[:11])
    if source_ref is None:
        for sfx in suffix_candidates(info.suffix):
            source_ref = style_color_suffix_to_source.get((source_style, info.color_code, sfx))
            if source_ref:
                break
    if source_ref is None:
        source_ref = style_color_to_source.get((source_style, info.color_code))
    return source_ref


def read_source_field(
    source_index: Dict[str, Any],
    source_sku: str,
    field_candidates: List[str],
) -> str:
    source_rows: Dict[str, List[Any]] = source_index["source_rows"]
    source_file_by_sku: Dict[str, str] = source_index["source_file_by_sku"]
    source_header_map_by_file: Dict[str, Dict[str, List[int]]] = source_index["source_header_map_by_file"]

    row_values = source_rows.get(source_sku, [])
    source_file = source_file_by_sku.get(source_sku, "")
    header_map = source_header_map_by_file.get(source_file, {})
    if not row_values or not header_map:
        return ""

    for field in field_candidates:
        key = excel_processor._normalize_header(field)
        cols = header_map.get(key, [])
        for col_idx in cols:
            if 0 < col_idx <= len(row_values):
                text = normalize_text(row_values[col_idx - 1])
                if text:
                    return text
    return ""


def build_source_index_for_store(template_type: str, data_files: List[str]) -> Dict[str, Any]:
    preferred_store = TEMPLATE_TO_STORE.get(template_type, "EP")
    indexes_by_store = excel_processor._build_source_indexes_cached(data_files, preferred_store)
    store_order = excel_processor._build_store_search_order(preferred_store)
    available_store_order = [store for store in store_order if store in indexes_by_store]
    if not available_store_order and indexes_by_store:
        available_store_order = sorted(indexes_by_store.keys())
    return excel_processor._merge_source_indexes(indexes_by_store, available_store_order)


def collect_store_result(template_type: str, generated_skus: List[str]) -> StoreRunResult:
    expected_store = TEMPLATE_TO_STORE.get(template_type, "EP")
    filenames = resolve_filenames(template_type)
    if not filenames:
        raise RuntimeError(f"{template_type} 未找到可用源文件")

    # 每次跑店铺用空缓存，确保可追踪索引库访问。
    excel_processor._price_map_cache.clear()
    excel_processor._asin_map_cache.clear()
    excel_processor._source_index_cache.clear()

    db_accesses: List[str] = []
    original_connect = excel_processor._connect_index_db

    def tracked_connect(store_prefix: str):
        db_accesses.append(normalize_text(store_prefix).upper() or "EP")
        return original_connect(store_prefix)

    excel_processor._connect_index_db = tracked_connect  # type: ignore[method-assign]
    try:
        output_filename, processed_count = excel_processor.process_excel(
            template_type=template_type,
            filenames=filenames,
            selected_prefixes=PREFIXES,
            generated_skus=generated_skus,
            target_color=None,
            target_size=None,
            processing_mode="add-code",
            follow_sell_mode=False,
        )
    finally:
        excel_processor._connect_index_db = original_connect  # type: ignore[method-assign]

    output_path = RESULTS_DIR / output_filename
    if not output_path.exists():
        raise RuntimeError(f"未找到输出文件: {output_path}")

    rows = read_output_rows(output_path)
    source_index = build_source_index_for_store(template_type, [f for f in filenames if not f.lower().endswith(".txt")])
    source_file_by_sku: Dict[str, str] = source_index["source_file_by_sku"]

    anomalies: List[str] = []
    allowed_source_files = set(STORE_CONFIGS.get(template_type, {}).get("source_files", []))

    for row in rows:
        source_sku = resolve_source_for_add_code(row.sku, source_index)
        row.source_sku = source_sku or ""
        if source_sku:
            row.source_file = source_file_by_sku.get(source_sku, "")
            row.expected_brand_name = read_source_field(
                source_index,
                source_sku,
                ["Brand Name", "Brand"],
            )
            row.expected_product_name = read_source_field(
                source_index,
                source_sku,
                ["Product Name", "Item Name"],
            )
            if row.expected_product_name:
                row.expected_product_name = align_product_name_size(row.expected_product_name, row.sku)
            if row.source_file and allowed_source_files and row.source_file not in allowed_source_files:
                anomalies.append(
                    f"{template_type} SKU={row.sku} 来源文件异常: {row.source_file} 不在 {sorted(allowed_source_files)}"
                )
            if row.expected_brand_name and normalize_text(row.brand_name) != normalize_text(row.expected_brand_name):
                anomalies.append(
                    f"{template_type} SKU={row.sku} Brand 不一致: output={row.brand_name} expected={row.expected_brand_name}"
                )
            if row.expected_product_name and normalize_text(row.product_name) != normalize_text(row.expected_product_name):
                anomalies.append(
                    f"{template_type} SKU={row.sku} Product 不一致: output={row.product_name} expected={row.expected_product_name}"
                )
        else:
            anomalies.append(f"{template_type} SKU={row.sku} 无法回溯来源 source_sku")

    accessed_unique = sorted(set(db_accesses))
    unexpected_db = [store for store in accessed_unique if store != expected_store]
    if not db_accesses:
        anomalies.append(f"{template_type} 未记录到索引库访问（可能完全命中内存缓存）")
    if unexpected_db:
        anomalies.append(
            f"{template_type} 索引库访问异常: expected={expected_store}, actual={accessed_unique}"
        )

    return StoreRunResult(
        template_type=template_type,
        expected_store=expected_store,
        filenames=filenames,
        output_filename=output_filename,
        processed_count=processed_count,
        db_accesses=db_accesses,
        rows=rows,
        anomalies=anomalies,
    )


def detect_cross_store_mismatch(results: List[StoreRunResult]) -> List[str]:
    issues: List[str] = []
    expected_by_store: Dict[str, Dict[str, Tuple[str, str]]] = {}
    output_by_store: Dict[str, Dict[str, Tuple[str, str]]] = {}

    for result in results:
        expected_map: Dict[str, Tuple[str, str]] = {}
        output_map: Dict[str, Tuple[str, str]] = {}
        for row in result.rows:
            expected_map[row.sku] = (normalize_text(row.expected_brand_name), normalize_text(row.expected_product_name))
            output_map[row.sku] = (normalize_text(row.brand_name), normalize_text(row.product_name))
        expected_by_store[result.template_type] = expected_map
        output_by_store[result.template_type] = output_map

    for result in results:
        store_name = result.template_type
        for row in result.rows:
            output_pair = output_by_store[store_name].get(row.sku, ("", ""))
            own_expected_pair = expected_by_store[store_name].get(row.sku, ("", ""))
            if not any(own_expected_pair):
                continue
            if output_pair == own_expected_pair:
                continue
            for other in results:
                if other.template_type == store_name:
                    continue
                other_expected = expected_by_store[other.template_type].get(row.sku, ("", ""))
                if any(other_expected) and output_pair == other_expected:
                    issues.append(
                        f"{store_name} SKU={row.sku} 输出值与 {other.template_type} 源值一致，疑似跨店铺串值"
                    )
    return issues


def print_store_table(result: StoreRunResult) -> None:
    print(f"\n=== {result.template_type} (expected store={result.expected_store}) ===")
    print(f"source files: {result.filenames}")
    print(f"output file : {result.output_filename}")
    print(f"processed   : {result.processed_count}")
    print(f"db accessed : {sorted(set(result.db_accesses))} (raw calls={len(result.db_accesses)})")

    print("| SKU | Brand Name | Product Name(前80字符) | 数据来源文件 |")
    print("|---|---|---|---|")
    for row in result.rows:
        print(
            f"| {row.sku} | {row.brand_name} | {product_preview(row.product_name)} | {row.source_file or '-'} |"
        )

    if result.anomalies:
        print("异常:")
        for issue in result.anomalies:
            print(f"- {issue}")
    else:
        print("异常: 无")


def main() -> int:
    generated_skus = build_generated_skus(PREFIXES, COLORS, SIZES)
    print("测试参数:")
    print(f"- prefixes: {PREFIXES}")
    print(f"- colors  : {COLORS}")
    print(f"- sizes   : {SIZES}")
    print(f"- generated_skus count: {len(generated_skus)}")

    results: List[StoreRunResult] = []
    for template_type in TEMPLATE_TYPES:
        result = collect_store_result(template_type, generated_skus)
        results.append(result)
        print_store_table(result)

    cross_store_issues = detect_cross_store_mismatch(results)
    print("\n=== 汇总结论 ===")
    if cross_store_issues:
        print("检测到疑似跨店铺串值:")
        for issue in cross_store_issues:
            print(f"- {issue}")
    else:
        print("未检测到跨店铺串值")

    any_anomaly = any(result.anomalies for result in results) or bool(cross_store_issues)
    return 1 if any_anomaly else 0


if __name__ == "__main__":
    raise SystemExit(main())
