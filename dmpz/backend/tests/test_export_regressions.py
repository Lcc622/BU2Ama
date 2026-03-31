from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.config import STORE_CONFIGS
import app.core.excel_processor as excel_processor_module
from app.core.excel_processor import ExcelProcessor
from app.core.follow_sell_processor import FollowSellProcessor


def _resolve_source_files(template_type: str) -> list[str]:
    source_files: list[str] = []
    config = STORE_CONFIGS[template_type]
    for filename in config.get("source_files", []):
        path = excel_processor_module.UPLOADS_DIR / filename
        if path.exists():
            source_files.append(filename)
    listing_report = str(config.get("listing_report", "")).strip()
    if listing_report and (excel_processor_module.UPLOADS_DIR / listing_report).exists():
        source_files.append(listing_report)
    return source_files


def _header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        for row_idx in (2, 3):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            key = str(value).strip().lower()
            if key and key not in mapping:
                mapping[key] = col_idx
    return mapping


def _read_export_rows(output_path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    try:
        ws = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
        headers = _header_map(ws)
        required = {
            "seller sku": headers["seller sku"],
            "parent sku": headers["parent sku"],
            "variation theme": headers.get("variation theme"),
            "colour map": headers.get("colour map"),
        }
        rows: list[dict[str, str]] = []
        for row_idx in range(4, ws.max_row + 1):
            sku = ws.cell(row=row_idx, column=required["seller sku"]).value
            if sku is None or str(sku).strip() == "":
                continue
            rows.append(
                {
                    "seller_sku": str(sku).strip(),
                    "parent_sku": str(ws.cell(row=row_idx, column=required["parent sku"]).value or "").strip(),
                    "variation_theme": (
                        str(ws.cell(row=row_idx, column=required["variation theme"]).value or "").strip()
                        if required["variation theme"]
                        else ""
                    ),
                    "colour_map": (
                        str(ws.cell(row=row_idx, column=required["colour map"]).value or "").strip()
                        if required["colour map"]
                        else ""
                    ),
                }
            )
        return rows
    finally:
        workbook.close()


def _read_single_cell_by_header(output_path: Path, row_idx: int, header_name: str) -> str:
    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    try:
        ws = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
        headers = _header_map(ws)
        col_idx = headers[header_name.strip().lower()]
        value = ws.cell(row=row_idx, column=col_idx).value
        return str(value).strip() if value is not None else ""
    finally:
        workbook.close()


def _read_rows_by_headers(output_path: Path, header_names: list[str]) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    try:
        ws = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
        headers = _header_map(ws)
        aliases = {
            "Colour Map": ["colour map", "color map"],
            "Color": ["color", "color_name"],
            "Item Length Description": ["item length description", "item_length_description"],
        }
        header_cols = {}
        for name in header_names:
            candidates = aliases.get(name, [name.strip().lower()])
            if name not in aliases:
                candidates = [name.strip().lower()]
            for candidate in candidates:
                if candidate in headers:
                    header_cols[name] = headers[candidate]
                    break
            else:
                raise KeyError(name.strip().lower())
        seller_sku_col = headers["seller sku"]
        rows: list[dict[str, str]] = []
        for row_idx in range(4, ws.max_row + 1):
            sku = ws.cell(row=row_idx, column=seller_sku_col).value
            if sku is None or str(sku).strip() == "":
                continue
            row_data = {"seller sku": str(sku).strip()}
            for name, col_idx in header_cols.items():
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[name] = str(value).strip() if value is not None else ""
            rows.append(row_data)
        return rows
    finally:
        workbook.close()


@pytest.fixture
def isolated_results_dir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    monkeypatch.setattr(excel_processor_module, "RESULTS_DIR", tmp_path)
    return tmp_path


def test_follow_sell_export_regression_ee00733wh(isolated_results_dir: Path) -> None:
    processor = FollowSellProcessor()
    query_result = processor.find_sizes_for_skc("EE00733WH", "EPUS")

    assert query_result["success"] is True
    assert query_result["old_style"] == "EE00736"

    generated_skus = [item["sku"] for item in query_result["sizes"]]
    assert generated_skus == [
        "EE00733WH04-USA",
        "EE00733WH06-USA",
        "EE00733WH08-USA",
        "EE00733WH10-USA",
        "EE00733WH12-USA",
        "EE00733WH14-USA",
        "EE00733WH16-USA",
        "EE00733WH18-USA",
        "EE00733WH20-USA",
        "EE00733WH22-USA",
        "EE00733WH24-USA",
        "EE00733WH26-USA",
    ]

    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="EPUS",
        filenames=_resolve_source_files("EPUS"),
        selected_prefixes=["EE00733"],
        generated_skus=generated_skus,
        target_color=None,
        target_size=None,
        source_style_map={"EE00733": "EE00736"},
        clear_image_urls=True,
        follow_sell_mode=True,
    )

    assert processed_count == 12

    rows = _read_export_rows(isolated_results_dir / output_filename)
    assert len(rows) == 12
    assert [row["seller_sku"] for row in rows] == generated_skus
    assert {row["parent_sku"] for row in rows} == {"EP00734-USB"}
    assert {row["variation_theme"] for row in rows} == {"SizeName-ColorName"}
    assert {row["colour_map"] for row in rows} == {"White"}


def test_add_color_export_regression_ee00466_bd(isolated_results_dir: Path) -> None:
    generated_skus = [
        f"EE00466BD{size:02d}"
        for size in range(4, 27, 2)
    ]

    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="EPUS",
        filenames=_resolve_source_files("EPUS"),
        selected_prefixes=["EE00466"],
        generated_skus=generated_skus,
        target_color=None,
        target_size=None,
        processing_mode="add-color",
    )

    assert processed_count == 12

    rows = _read_export_rows(isolated_results_dir / output_filename)
    assert len(rows) == 12
    assert [row["seller_sku"] for row in rows] == [
        "EE00466BD04-USA",
        "EE00466BD06-USA",
        "EE00466BD08-USA",
        "EE00466BD10-USA",
        "EE00466BD12-USA",
        "EE00466BD14-PH",
        "EE00466BD16-PH",
        "EE00466BD18-PH",
        "EE00466BD20-PH",
        "EE00466BD22-PH",
        "EE00466BD24-PH",
        "EE00466BD26-PH",
    ]
    assert {row["parent_sku"] for row in rows} == {"EP00461-USC"}
    assert {row["variation_theme"] for row in rows} == {"SizeName-ColorName"}
    assert {row["colour_map"] for row in rows} == {"Red"}


def test_sizecolor_export_keeps_source_color_fields(isolated_results_dir: Path) -> None:
    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="EPUS",
        filenames=_resolve_source_files("EPUS"),
        selected_prefixes=["EZ07707"],
        generated_skus=["EZ07707BD04"],
        target_color=None,
        target_size=None,
        processing_mode="add-color",
    )

    assert processed_count == 1

    rows = _read_export_rows(isolated_results_dir / output_filename)
    assert rows == [
        {
            "seller_sku": "EZ07707BD04-USA",
            "parent_sku": "EZ07707-USA",
            "variation_theme": "SizeColor",
            "colour_map": "Burgundy",
        }
    ]


def test_add_code_dual_listing_expands_both_suffixes(isolated_results_dir: Path) -> None:
    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="EPUS",
        filenames=_resolve_source_files("EPUS"),
        selected_prefixes=["ES0128B"],
        generated_skus=["ES0128BFL02"],
        target_color=None,
        target_size=None,
        processing_mode="add-code",
    )

    assert processed_count == 2

    rows = _read_export_rows(isolated_results_dir / output_filename)
    assert rows == [
        {
            "seller_sku": "ES0128BFL02-USA",
            "parent_sku": "ES0128B",
            "variation_theme": "SizeName-ColorName",
            "colour_map": "Multicolor",
        },
        {
            "seller_sku": "ES0128BFL02-PH",
            "parent_sku": "ES0128B-PH",
            "variation_theme": "SizeName-ColorName",
            "colour_map": "Multicolor",
        },
    ]


def test_add_code_uses_source_images_for_ep_mg02(isolated_results_dir: Path) -> None:
    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="EPUS",
        filenames=_resolve_source_files("EPUS"),
        selected_prefixes=["ES01955"],
        generated_skus=["ES01955MG02"],
        target_color=None,
        target_size=None,
        processing_mode="add-code",
    )

    assert processed_count == 1

    output_path = isolated_results_dir / output_filename
    rows = _read_export_rows(output_path)
    assert rows == [
        {
            "seller_sku": "ES01955MG02-USA",
            "parent_sku": "ES01955-USA",
            "variation_theme": "SizeName-ColorName",
            "colour_map": "Green",
        }
    ]
    assert _read_single_cell_by_header(output_path, 4, "Main Image URL") == (
        "https://m.media-amazon.com/images/I/61Q+RPJvTEL.jpg"
    )
    assert _read_single_cell_by_header(output_path, 4, "Swatch Image URL") == (
        "https://m.media-amazon.com/images/I/61Q+RPJvTEL.jpg"
    )


def test_add_code_uses_matched_source_images_for_dual_listing_ep_fl02(isolated_results_dir: Path) -> None:
    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="EPUS",
        filenames=_resolve_source_files("EPUS"),
        selected_prefixes=["ES0128B"],
        generated_skus=["ES0128BFL02"],
        target_color=None,
        target_size=None,
        processing_mode="add-code",
    )

    assert processed_count == 2

    output_path = isolated_results_dir / output_filename
    assert _read_single_cell_by_header(output_path, 4, "Main Image URL") == (
        "https://m.media-amazon.com/images/I/61+-33sNLCL.jpg"
    )
    assert _read_single_cell_by_header(output_path, 5, "Main Image URL") == (
        "https://m.media-amazon.com/images/I/61wQyZftB+L.jpg"
    )


def test_add_code_source_matching_rejects_cross_color_fallback() -> None:
    excel_processor = ExcelProcessor()
    info = excel_processor.parse_sku("EE02665GD02-USA")

    assert info is not None

    source_ref = excel_processor._resolve_add_code_source_ref(
        requested_sku="EE02665GD02-USA",
        info=info,
        source_style="EE02665",
        sku_to_source={},
        sku_base_to_source={"EE02665GD02": "EE02665KG02-USA"},
        style_color_suffix_to_source={("EE02665", "GD", "-USA"): "EE02665GD04-USA"},
        style_color_to_source={("EE02665", "GD"): "EE02665GD14-USA"},
        suffixes=["-USA", "-PH", ""],
    )

    assert source_ref == "EE02665GD04-USA"


def test_add_color_respects_explicit_mode_for_new_dm_color(isolated_results_dir: Path) -> None:
    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="DaMaUS",
        filenames=_resolve_source_files("DaMaUS"),
        selected_prefixes=["ES01955"],
        generated_skus=[
            "ES01955NT14",
            "ES01955NT16",
            "ES01955NT18",
            "ES01955NT20",
            "ES01955NT22",
            "ES01955NT24",
            "ES01955NT26",
        ],
        target_color=None,
        target_size=None,
        processing_mode="add-color",
    )

    assert processed_count == 7

    rows = _read_export_rows(isolated_results_dir / output_filename)
    assert [row["seller_sku"] for row in rows] == [
        "ES01955NT14-PLPH",
        "ES01955NT16-PLPH",
        "ES01955NT18-PLPH",
        "ES01955NT20-PLPH",
        "ES01955NT22-PLPH",
        "ES01955NT24-PLPH",
        "ES01955NT26-PLPH",
    ]


def test_add_color_dm_pe_export_uses_fixed_theme_color_and_quantity(isolated_results_dir: Path) -> None:
    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="DaMaUS",
        filenames=_resolve_source_files("DaMaUS"),
        selected_prefixes=["ES01955"],
        generated_skus=[
            "ES01955PE14",
            "ES01955PE16",
            "ES01955PE18",
            "ES01955PE20",
            "ES01955PE22",
            "ES01955PE24",
            "ES01955PE26",
        ],
        target_color=None,
        target_size=None,
        processing_mode="add-color",
    )

    assert processed_count == 7

    rows = _read_rows_by_headers(
        isolated_results_dir / output_filename,
        [
            "Product Name",
            "Variation Theme",
            "Colour Map",
            "Color",
            "Quantity",
            "Item Length Description",
        ],
    )
    assert rows == [
        {
            "seller sku": "ES01955PE14-PLPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US14",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
        },
        {
            "seller sku": "ES01955PE16-PLPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US16",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
        },
        {
            "seller sku": "ES01955PE18-PLPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US18",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
        },
        {
            "seller sku": "ES01955PE20-PLPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US20",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
        },
        {
            "seller sku": "ES01955PE22-PLPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US22",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
        },
        {
            "seller sku": "ES01955PE24-PLPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US24",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
        },
        {
            "seller sku": "ES01955PE26-PLPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US26",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
        },
    ]


def test_add_color_pz_pe_export_uses_consistent_title_and_fixed_fields(isolated_results_dir: Path) -> None:
    excel_processor = ExcelProcessor()
    output_filename, processed_count = excel_processor.process_excel(
        template_type="PZUS",
        filenames=_resolve_source_files("PZUS"),
        selected_prefixes=["ES01955"],
        generated_skus=[
            "ES01955PE14",
            "ES01955PE16",
            "ES01955PE18",
            "ES01955PE20",
            "ES01955PE22",
            "ES01955PE24",
            "ES01955PE26",
        ],
        target_color=None,
        target_size=None,
        processing_mode="add-color",
    )

    assert processed_count == 7

    rows = _read_rows_by_headers(
        isolated_results_dir / output_filename,
        [
            "Product Name",
            "Variation Theme",
            "Colour Map",
            "Color",
            "Quantity",
            "Item Length Description",
            "bullet_point1",
            "bullet_point2",
            "bullet_point3",
            "bullet_point4",
            "bullet_point5",
            "embellishment_feature1",
            ],
    )
    assert rows == [
        {
            "seller sku": "ES01955PE14-DAPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US14",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
            "bullet_point1": "Size Attention: Please check the size chart before ordering. If your measurements vary, use the smallest as a guide. For example, if your bust is size 8 and waist is size 10, choose size 8 based on your bust",
            "bullet_point2": "Features: Floor length, sleeveless, chiffon, crafted from no stretch fabric, not padded, with lining, concealed zipper up the back, a line, round neck, pleated, hollow out, sequin waist",
            "bullet_point3": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point4": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point5": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "embellishment_feature1": "Chiffon",
        },
        {
            "seller sku": "ES01955PE16-DAPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US16",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
            "bullet_point1": "Size Attention: Please check the size chart before ordering. If your measurements vary, use the smallest as a guide. For example, if your bust is size 8 and waist is size 10, choose size 8 based on your bust",
            "bullet_point2": "Features: Floor length, sleeveless, chiffon, crafted from no stretch fabric, not padded, with lining, concealed zipper up the back, a line, round neck, pleated, hollow out, sequin waist",
            "bullet_point3": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point4": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point5": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "embellishment_feature1": "Chiffon",
        },
        {
            "seller sku": "ES01955PE18-DAPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US18",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
            "bullet_point1": "Size Attention: Please check the size chart before ordering. If your measurements vary, use the smallest as a guide. For example, if your bust is size 8 and waist is size 10, choose size 8 based on your bust",
            "bullet_point2": "Features: Floor length, sleeveless, chiffon, crafted from no stretch fabric, not padded, with lining, concealed zipper up the back, a line, round neck, pleated, hollow out, sequin waist",
            "bullet_point3": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point4": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point5": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "embellishment_feature1": "Chiffon",
        },
        {
            "seller sku": "ES01955PE20-DAPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US20",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
            "bullet_point1": "Size Attention: Please check the size chart before ordering. If your measurements vary, use the smallest as a guide. For example, if your bust is size 8 and waist is size 10, choose size 8 based on your bust",
            "bullet_point2": "Features: Floor length, sleeveless, chiffon, crafted from no stretch fabric, not padded, with lining, concealed zipper up the back, a line, round neck, pleated, hollow out, sequin waist",
            "bullet_point3": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point4": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point5": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "embellishment_feature1": "Chiffon",
        },
        {
            "seller sku": "ES01955PE22-DAPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US22",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
            "bullet_point1": "Size Attention: Please check the size chart before ordering. If your measurements vary, use the smallest as a guide. For example, if your bust is size 8 and waist is size 10, choose size 8 based on your bust",
            "bullet_point2": "Features: Floor length, sleeveless, chiffon, crafted from no stretch fabric, not padded, with lining, concealed zipper up the back, a line, round neck, pleated, hollow out, sequin waist",
            "bullet_point3": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point4": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point5": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "embellishment_feature1": "Chiffon",
        },
        {
            "seller sku": "ES01955PE24-DAPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US24",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
            "bullet_point1": "Size Attention: Please check the size chart before ordering. If your measurements vary, use the smallest as a guide. For example, if your bust is size 8 and waist is size 10, choose size 8 based on your bust",
            "bullet_point2": "Features: Floor length, sleeveless, chiffon, crafted from no stretch fabric, not padded, with lining, concealed zipper up the back, a line, round neck, pleated, hollow out, sequin waist",
            "bullet_point3": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point4": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point5": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "embellishment_feature1": "Chiffon",
        },
        {
            "seller sku": "ES01955PE26-DAPH",
            "Product Name": "Ever-Pretty Plus Women's Round Neck Sleeveless Empire Waist Maxi Plus Size Formal Gowns Bridesmaid Dresses Peach US26",
            "Variation Theme": "SizeName-ColorName",
            "Colour Map": "Orange",
            "Color": "Peach",
            "Quantity": "0",
            "Item Length Description": "Floor Length",
            "bullet_point1": "Size Attention: Please check the size chart before ordering. If your measurements vary, use the smallest as a guide. For example, if your bust is size 8 and waist is size 10, choose size 8 based on your bust",
            "bullet_point2": "Features: Floor length, sleeveless, chiffon, crafted from no stretch fabric, not padded, with lining, concealed zipper up the back, a line, round neck, pleated, hollow out, sequin waist",
            "bullet_point3": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point4": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "bullet_point5": "Highlights: This plus size bridesmaid dress features a chic round neck with a stylish hollow-out detail and a sequin-accented waist that highlights your curves beautifully — elegant, flattering, and perfect for any special occasion",
            "embellishment_feature1": "Chiffon",
        },
    ]

    price_rows = _read_rows_by_headers(
        isolated_results_dir / output_filename,
        ["standard_price", "list_price"],
    )
    assert len(price_rows) == 7
    assert all(row["standard_price"] == "83.99" for row in price_rows)
    assert all(row["list_price"] == "93.99" for row in price_rows)
