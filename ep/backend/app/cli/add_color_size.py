#!/usr/bin/env python3
"""加色加码 CLI 入口。"""

from __future__ import annotations

import argparse
import contextlib
import io
import re
import sys
from pathlib import Path

from utils import STORE_TO_TEMPLATE, bootstrap_app, normalize_store, print_result

bootstrap_app()

from app.config import RESULTS_DIR, STORE_CONFIGS, UPLOADS_DIR
from app.core.export_history import export_history
import openpyxl


def count_template_rows(template_path: Path) -> int:
    """Count non-empty SKU rows in the template workbook."""
    workbook = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
        total = 0
        for row in worksheet.iter_rows(min_row=7, min_col=3, max_col=3, values_only=True):
            if row and row[0]:
                total += 1
        return total
    finally:
        workbook.close()


def resolve_source_files(store: str, provided_sources: list[str] | None) -> tuple[list[str], list[str]]:
    """Resolve source files for the selected store."""
    warnings: list[str] = []
    if provided_sources:
        source_files = [str(Path(name).name) for name in provided_sources if str(name).strip()]
    else:
        source_files = []
        config = STORE_CONFIGS.get({"EP": "EPUS", "DM": "DaMaUS", "PZ": "PZUS"}[store], {})
        for file_name in config.get("source_files", []):
            if (UPLOADS_DIR / file_name).exists():
                source_files.append(file_name)
        if not source_files:
            from app.core.excel_processor import ExcelProcessor
            processor = ExcelProcessor()
            scanned = processor._scan_store_data_files().get(store, [])  # noqa: SLF001
            source_files = list(scanned)

    existing: list[str] = []
    for source_file in source_files:
        if (UPLOADS_DIR / source_file).exists():
            existing.append(source_file)
        else:
            warnings.append(f"Source file not found and skipped: {source_file}")
    return existing, warnings


def split_values(values: list[str] | None) -> list[str]:
    """Split comma or whitespace separated CLI values."""
    items: list[str] = []
    for raw in values or []:
        for part in re.split(r"[,，、;；\s]+", str(raw).strip()):
            value = part.strip()
            if value:
                items.append(value)
    return items


def normalize_prefixes(values: list[str] | None) -> list[str]:
    """Normalize one or more product prefixes."""
    prefixes: list[str] = []
    seen: set[str] = set()
    for raw in split_values(values):
        prefix = raw.strip().upper()
        if not prefix or prefix in seen:
            continue
        seen.add(prefix)
        prefixes.append(prefix)
    return prefixes


def parse_color_codes(values: list[str] | None) -> list[str]:
    """Normalize two-letter color codes from CLI input."""
    colors: list[str] = []
    seen: set[str] = set()
    for raw in split_values(values):
        color = raw.strip().upper()
        if len(color) != 2 or color in seen:
            continue
        seen.add(color)
        colors.append(color)
    return colors


def parse_sizes(values: list[str] | None) -> list[int]:
    """Parse size values and preserve user ordering."""
    sizes: list[int] = []
    seen: set[int] = set()
    for raw in split_values(values):
        try:
            size = int(raw)
        except ValueError:
            continue
        if size in seen:
            continue
        seen.add(size)
        sizes.append(size)
    return sizes


def build_generated_skus(
    prefixes: list[str],
    mode: str,
    colors: list[str],
    start_size: str | None,
    end_size: str | None,
    size_step: int,
    explicit_sizes: list[str] | None,
) -> list[str]:
    """Generate target SKUs using the same logic as the web UI."""
    if not prefixes:
        return []
    if not colors:
        return []

    parsed_size_list = parse_sizes(explicit_sizes or ([start_size] if start_size else []))

    sizes: list[int] = []
    if mode == "add-color":
        try:
            start = int(str(start_size or "").strip())
            end = int(str(end_size or "").strip())
        except ValueError:
            start = -1
            end = -1
        if start >= 0 and end >= 0 and start <= end and size_step > 0:
            sizes = list(range(start, end + 1, size_step))
        else:
            sizes = parsed_size_list
    else:
        sizes = parsed_size_list

    if not sizes:
        return []

    generated: list[str] = []
    seen: set[str] = set()
    for prefix in prefixes:
        for color in colors:
            for size in sizes:
                sku = f"{prefix}{color}{size:02d}"
                if sku in seen:
                    continue
                seen.add(sku)
                generated.append(sku)
    return generated


def resolve_process_source_files(store: str, provided_sources: list[str] | None) -> tuple[list[str], list[str]]:
    """Resolve data files for process_excel web-style flow."""
    warnings: list[str] = []
    template_type = STORE_TO_TEMPLATE[store]
    config = STORE_CONFIGS[template_type]

    source_files, warnings = resolve_source_files(store, provided_sources)
    listing_report = str(config.get("listing_report", "")).strip()
    if listing_report:
        report_path = UPLOADS_DIR / listing_report
        if report_path.exists():
            source_files.append(listing_report)
        else:
            warnings.append(f"Listing report not found and skipped: {listing_report}")
    return source_files, warnings


def record_export_history(
    *,
    module: str,
    template_type: str,
    input_data: dict[str, object],
    filename: str,
    processed_count: int,
) -> None:
    """Store export history the same way as the API layer."""
    file_path = RESULTS_DIR / filename
    if not file_path.exists():
        return
    export_history.add_record(
        module=module,
        template_type=template_type,
        input_data=input_data,
        filename=filename,
        file_size=int(file_path.stat().st_size),
        processed_count=processed_count,
        status="success",
    )
    export_history.cleanup(retention_days=90, max_records=1000)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="加色加码处理")
    parser.add_argument("--template", help="模板文件名或路径（兼容旧模式）")
    parser.add_argument("--store", required=True, choices=["EP", "DM", "PZ"], help="店铺代码")
    parser.add_argument("--sources", nargs="+", help="源文件列表")
    parser.add_argument("--price-report", help="价格报告文件名或路径")
    parser.add_argument("--prefix", action="append", default=[], help="产品前缀，可重复或逗号分隔")
    parser.add_argument("--mode", choices=["add-color", "add-code"], default="add-color", help="网页同款处理模式")
    parser.add_argument("--colors", action="append", default=[], help="颜色代码，可重复或逗号分隔")
    parser.add_argument("--start-size", help="起始尺码")
    parser.add_argument("--end-size", help="结束尺码")
    parser.add_argument("--size-step", type=int, default=2, help="尺码步长")
    parser.add_argument("--sizes", action="append", default=[], help="尺码列表，可重复或逗号分隔")
    parser.add_argument("--json", action="store_true", help="输出 JSON")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    result = {
        "success": False,
        "output_file": None,
        "processed_count": 0,
        "skipped_count": 0,
        "errors": [],
        "warnings": [],
    }

    try:
        store = normalize_store(args.store)
        prefixes = normalize_prefixes(args.prefix)

        if prefixes:
            template_type = STORE_TO_TEMPLATE[store]
            source_files, warnings = resolve_process_source_files(store, args.sources)
            result["warnings"].extend(warnings)
            if not source_files:
                raise FileNotFoundError(f"No source files found for store {store}")

            generated_skus = build_generated_skus(
                prefixes=prefixes,
                mode=args.mode,
                colors=parse_color_codes(args.colors),
                start_size=args.start_size,
                end_size=args.end_size,
                size_step=max(1, int(args.size_step or 1)),
                explicit_sizes=args.sizes,
            )
            if not generated_skus:
                raise ValueError("No valid target SKUs generated. Check prefix, colors, and size inputs.")

            with contextlib.redirect_stdout(io.StringIO()):
                from app.core.excel_processor import ExcelProcessor
                processor = ExcelProcessor()
                output_filename, processed_count = processor.process_excel(
                    template_type=template_type,
                    filenames=source_files,
                    selected_prefixes=prefixes,
                    generated_skus=generated_skus,
                    target_color=None,
                    target_size=None,
                    processing_mode=args.mode,
                )

            output_path = RESULTS_DIR / output_filename
            record_export_history(
                module=args.mode,
                template_type=template_type,
                input_data={
                    "prefixes": prefixes,
                    "colors": sorted({sku[7:9] for sku in generated_skus if len(sku) >= 9}),
                    "sizes": sorted({sku[9:11] for sku in generated_skus if len(sku) >= 11}),
                    "mode": args.mode,
                },
                filename=output_filename,
                processed_count=int(processed_count),
            )
            result.update(
                {
                    "success": True,
                    "mode": args.mode,
                    "template_type": template_type,
                    "selected_prefixes": prefixes,
                    "source_files": source_files,
                    "output_file": str(output_path if output_path.exists() else output_filename),
                    "processed_count": int(processed_count),
                    "skipped_count": max(0, len(generated_skus) - int(processed_count)),
                    "total_skus": len(generated_skus),
                    "generated_skus_preview": generated_skus[:20],
                }
            )
        else:
            if not args.template:
                raise ValueError("Either --template or at least one --prefix is required.")

            template_name = Path(args.template).name
            template_path = UPLOADS_DIR / template_name
            if not template_path.exists():
                raise FileNotFoundError(f"Template file not found: {template_name}")

            source_files, warnings = resolve_source_files(store, args.sources)
            result["warnings"].extend(warnings)
            if not source_files:
                raise FileNotFoundError(f"No source files found for store {store}")

            price_report = Path(args.price_report).name if args.price_report else None
            if price_report and not (UPLOADS_DIR / price_report).exists():
                raise FileNotFoundError(f"Price report file not found: {price_report}")

            total_rows = count_template_rows(template_path)

            with contextlib.redirect_stdout(io.StringIO()):
                from app.core.excel_processor import ExcelProcessor
                processor = ExcelProcessor()
                output_filename, processed_count = processor.process_excel_new(
                    template_filename=template_name,
                    source_filenames=source_files,
                    price_report_filename=price_report,
                )

            output_path = RESULTS_DIR / output_filename
            result.update(
                {
                    "success": True,
                    "output_file": str(output_path if output_path.exists() else output_filename),
                    "processed_count": int(processed_count),
                    "skipped_count": max(0, int(total_rows) - int(processed_count)),
                    "source_files": source_files,
                }
            )
        print_result(result, json_mode=args.json)
        return 0
    except Exception as exc:
        result["errors"].append(str(exc))
        print_result(result, json_mode=args.json)
        return 1


if __name__ == "__main__":
    sys.exit(main())
