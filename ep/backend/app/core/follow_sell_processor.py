"""
跟卖模块处理器
处理 SKC 输入，通过新老款映射查找尺码信息
"""
import sqlite3
import openpyxl
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from app.config import UPLOADS_DIR, STORE_CONFIGS

STORE_PREFIX_PATTERN = re.compile(r"^(EP|DM|DA|PZ)-\d", re.IGNORECASE)
STORE_PRIORITY = ("EP", "DM", "PZ")


class FollowSellProcessor:
    """跟卖处理器"""

    def __init__(self):
        self.new_to_old_mapping: Dict[str, str] = {}
        self._last_parse_error = ""
        self._source_field_cache: Dict[Tuple[str, str, Tuple[str, ...]], Optional[str]] = {}
        legacy_db_path = UPLOADS_DIR / "ep_index.db"
        if legacy_db_path.exists():
            print("警告: 检测到旧索引库 ep_index.db，请重新上传店铺数据源以重建 EP/DM/PZ 分店铺索引。")
        self._load_mapping()
        # 不在初始化时加载店铺索引，按店铺懒加载

    def _build_files_signature(self, filenames: List[str]) -> str:
        parts: List[str] = []
        for filename in sorted(filenames):
            path = UPLOADS_DIR / filename
            if not path.exists():
                continue
            stat = path.stat()
            parts.append(f"{filename}:{stat.st_mtime_ns}:{stat.st_size}")
        return "|".join(parts)

    def _normalize_store_prefix(self, store_prefix: Optional[str]) -> str:
        normalized = str(store_prefix or "").strip().upper()
        if normalized not in STORE_PRIORITY:
            normalized = "EP"
        return normalized

    def get_store_prefix(self, template_type: Optional[str]) -> str:
        text = str(template_type or "").strip().upper()
        if "PZ" in text:
            return "PZ"
        if "DM" in text or "DAMA" in text or text.startswith("DA"):
            return "DM"
        if "EP" in text:
            return "EP"
        return "EP"

    def get_template_type(self, store_prefix: Optional[str]) -> str:
        return "EPUS"

    def get_index_db_path(self, store_prefix: str) -> Path:
        normalized = self._normalize_store_prefix(store_prefix)
        return UPLOADS_DIR / f"ep_index_{normalized}.db"

    def _connect_db(self, store_prefix: str) -> sqlite3.Connection:
        conn = sqlite3.connect(self.get_index_db_path(store_prefix))
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        return conn

    def _ensure_index_tables(self, conn: sqlite3.Connection) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS ep_sku_index (
                store_prefix TEXT NOT NULL DEFAULT 'EP',
                style TEXT NOT NULL,
                sku TEXT NOT NULL,
                size TEXT NOT NULL,
                suffix TEXT NOT NULL,
                source_file TEXT NOT NULL,
                PRIMARY KEY (sku, source_file)
            )
            """
        )
        columns = {str(row[1]).lower() for row in conn.execute("PRAGMA table_info(ep_sku_index)").fetchall()}
        if "store_prefix" not in columns:
            conn.execute("ALTER TABLE ep_sku_index ADD COLUMN store_prefix TEXT NOT NULL DEFAULT 'EP'")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ep_style ON ep_sku_index(style)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ep_style_size ON ep_sku_index(style, size)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ep_store_style ON ep_sku_index(store_prefix, style)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ep_store_style_size ON ep_sku_index(store_prefix, style, size)")
        conn.commit()

    def _scan_store_data_files(self) -> Dict[str, List[str]]:
        files_by_store: Dict[str, List[str]] = {store: [] for store in STORE_PRIORITY}
        for path in sorted(UPLOADS_DIR.iterdir()):
            if not path.is_file():
                continue
            if path.suffix.lower() not in {".xlsm", ".xlsx"}:
                continue
            match = STORE_PREFIX_PATTERN.match(path.name.strip().upper())
            if not match:
                continue
            store_prefix = self._normalize_store_prefix(match.group(1))
            files_by_store[store_prefix].append(path.name)
        return {
            store: sorted({name for name in names})
            for store, names in files_by_store.items()
        }

    def _resolve_store_files(self, store_prefix: str) -> List[str]:
        normalized = self._normalize_store_prefix(store_prefix)
        scanned_files = self._scan_store_data_files()
        resolved = list(scanned_files.get(normalized, []))

        for template_type, config in STORE_CONFIGS.items():
            if self.get_store_prefix(template_type) != normalized:
                continue
            for file_name in config.get("source_files", []):
                file_text = str(file_name)
                if file_text not in resolved:
                    resolved.append(file_text)

        if resolved:
            return sorted(resolved)

        if normalized == "DM":
            return ["DA-0.xlsm", "DA-1.xlsm", "DA-2.xlsm", "DM-0.xlsm", "DM-1.xlsm", "DM-2.xlsm"]
        return [f"{normalized}-0.xlsm", f"{normalized}-1.xlsm", f"{normalized}-2.xlsm"]

    def _ensure_store_index(self, store_prefix: str) -> None:
        """确保指定店铺索引可用，文件变化时自动重建"""
        normalized_prefix = self._normalize_store_prefix(store_prefix)
        source_files = self._resolve_store_files(normalized_prefix)
        sku_pattern = re.compile(
            r"^(?P<style>[A-Z0-9]{7})(?P<color>[A-Z]{2})(?P<size>\d{2})(?P<suffix>-?[A-Z0-9]+)?$"
        )
        signature = self._build_files_signature(source_files)
        if not signature:
            return

        conn = self._connect_db(normalized_prefix)
        try:
            self._ensure_index_tables(conn)
            signature_key = f"signature:{normalized_prefix}"
            row = conn.execute("SELECT value FROM meta WHERE key=?", (signature_key,)).fetchone()
            current = row[0] if row else ""
            if current == signature:
                return

            print(f"{normalized_prefix} 文件有更新，正在重建 SQLite 索引...")
            conn.execute("DELETE FROM ep_sku_index WHERE store_prefix = ?", (normalized_prefix,))

            insert_sql = """
                INSERT OR REPLACE INTO ep_sku_index(store_prefix, style, sku, size, suffix, source_file)
                VALUES (?, ?, ?, ?, ?, ?)
            """
            total_filtered = 0

            for source_file in source_files:
                file_path = UPLOADS_DIR / source_file
                if not file_path.exists():
                    print(f"警告: 店铺文件不存在: {source_file}")
                    continue
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb['Template']
                    row_count = 0
                    filtered_count = 0
                    batch: List[Tuple[str, str, str, str, str, str]] = []
                    for row in ws.iter_rows(min_row=7, min_col=3, max_col=3, values_only=True):
                        sku_cell = row[0] if row else None
                        if not sku_cell:
                            continue
                        sku = str(sku_cell).strip().upper()
                        matched = sku_pattern.fullmatch(sku)
                        if not matched:
                            filtered_count += 1
                            continue
                        style = matched.group("style")
                        size = matched.group("size")
                        suffix = matched.group("suffix") or ""
                        batch.append((normalized_prefix, style, sku, size, suffix, source_file))
                        row_count += 1
                        if len(batch) >= 5000:
                            conn.executemany(insert_sql, batch)
                            conn.commit()
                            batch.clear()
                    if batch:
                        conn.executemany(insert_sql, batch)
                        conn.commit()
                    wb.close()
                    total_filtered += filtered_count
                    print(f"  {source_file} 索引完成，共 {row_count} 行，过滤无效 SKU {filtered_count} 行")
                except Exception as e:
                    print(f"加载 {source_file} 失败: {e}")
            print(f"{normalized_prefix} 索引重建完成，累计过滤无效 SKU {total_filtered} 行")

            conn.execute(
                "INSERT OR REPLACE INTO meta(key, value) VALUES(?, ?)",
                (signature_key, signature)
            )
            conn.commit()
        finally:
            conn.close()

    def _load_mapping(self):
        """加载新老款映射表"""
        mapping_file = UPLOADS_DIR / "新老款映射信息(1).xlsx"

        if not mapping_file.exists():
            print(f"警告: 新老款映射文件不存在: {mapping_file}")
            return

        try:
            wb = openpyxl.load_workbook(mapping_file, read_only=True, data_only=True)
            ws = wb.active

            # 从第2行开始读取（第1行是表头）
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] and row[1]:
                    old_style = str(row[0]).strip()
                    new_style = str(row[1]).strip()
                    self.new_to_old_mapping[new_style] = old_style

            wb.close()
            print(f"加载新老款映射: {len(self.new_to_old_mapping)} 条")

        except Exception as e:
            print(f"加载新老款映射失败: {e}")

    def _load_ep_data(self):
        """兼容旧接口：默认确保 EP 店铺索引"""
        self._ensure_store_index("EP")

    def parse_skc(self, skc: str) -> Optional[Tuple[str, str]]:
        """解析 SKC 格式

        Args:
            skc: SKC 字符串，格式为 7位style + 2位color，如 ES0128BDG

        Returns:
            (style, color_code) 或 None
        """
        skc = str(skc or "").strip().upper()
        self._last_parse_error = ""
        if not re.fullmatch(r"^[A-Z0-9]{7}[A-Z]{2}$", skc):
            self._last_parse_error = (
                f"SKC 格式错误: {skc or '<empty>'}，应匹配 ^[A-Z0-9]{{7}}[A-Z]{{2}}$（7位款号+2位大写字母颜色）"
            )
            return None

        style = skc[:7]  # 前7位
        color_code = skc[7:9]  # 后2位

        return style, color_code

    def _normalize_suffix(self, suffix: str) -> str:
        """按 SOP 统一 SKU 后缀。"""
        normalized = str(suffix or "").strip().upper()
        if not normalized:
            return "-USA"
        if not normalized.startswith("-"):
            normalized = f"-{normalized}"
        if normalized == "-":
            return "-USA"

        if normalized in {"-A", "-B", "-C", "-D"}:
            return "-USA"
        if normalized in {"-US", "-USB", "-USC", "-USD", "-USA"}:
            return "-USA"
        if normalized in {"-PH", "-PHB", "-PHC"}:
            return "-PH"
        return normalized

    def _generate_suffix_for_store(self, store_prefix: str, size: str) -> str:
        """根据店铺和尺码生成标准后缀。"""
        size_text = str(size or "")
        match = re.search(r"\d+", size_text)
        size_num = int(match.group()) if match else 0
        is_plus = size_num >= 14

        return "-PH" if is_plus else "-USA"

    def _extract_suffix_from_parent_sku(self, parent_sku: Optional[str]) -> str:
        """Extract and normalize suffix from a source parent SKU value."""
        text = str(parent_sku or "").strip().upper()
        if len(text) <= 7:
            return ""
        return self._normalize_suffix(text[7:])

    def _read_source_row_field(
        self,
        source_file: str,
        target_sku: str,
        field_names: List[str],
    ) -> Optional[str]:
        """Read a field value from a source workbook row identified by SKU."""
        cache_key = (
            str(source_file).strip(),
            str(target_sku).strip().upper(),
            tuple(str(name).strip().lower() for name in field_names),
        )
        if cache_key in self._source_field_cache:
            return self._source_field_cache[cache_key]

        file_path = UPLOADS_DIR / source_file
        if not file_path.exists():
            self._source_field_cache[cache_key] = None
            return None

        result: Optional[str] = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb['Template'] if 'Template' in wb.sheetnames else wb.active

            header_sku_col = 3 if str(ws.cell(row=4, column=3).value or '').strip().upper() == 'SKU' else 2
            header_map: Dict[str, List[int]] = {}
            for col_idx in range(1, ws.max_column + 1):
                for header_row in (4, 5):
                    value = ws.cell(row=header_row, column=col_idx).value
                    if not value:
                        continue
                    key = str(value).strip().lower()
                    if not key:
                        continue
                    header_map.setdefault(key, []).append(col_idx)

            normalized_target = str(target_sku).strip().upper()
            normalized_candidates = [str(name).strip().lower() for name in field_names if str(name).strip()]
            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) < header_sku_col:
                    continue
                sku_value = row[header_sku_col - 1]
                if str(sku_value or "").strip().upper() != normalized_target:
                    continue
                for field_name in normalized_candidates:
                    for col_idx in header_map.get(field_name, []):
                        if 0 < col_idx <= len(row):
                            value = row[col_idx - 1]
                            if value is not None and str(value).strip():
                                result = str(value).strip()
                                break
                    if result is not None:
                        break
                break
            wb.close()
        except Exception as exc:
            print(f"读取源字段失败 {source_file} {target_sku}: {exc}")
            result = None

        self._source_field_cache[cache_key] = result
        return result

    def find_sizes_for_skc(self, skc: str, template_type: str = "EPUS") -> Dict:
        """根据 SKC 查找所有尺码信息

        Args:
            skc: SKC 字符串，如 ES0128BDG

        Returns:
            {
                'success': bool,
                'skc': str,
                'new_style': str,
                'old_style': str,
                'color_code': str,
                'sizes': [{'size': str, 'sku': str, 'suffix': str}],
                'message': str
            }
        """
        store_prefix = self._normalize_store_prefix(self.get_store_prefix(template_type))
        self._ensure_store_index(store_prefix)

        result = {
            'success': False,
            'skc': skc,
            'new_style': '',
            'old_style': '',
            'color_code': '',
            'sizes': [],
            'message': ''
        }

        # 1. 解析 SKC
        parsed = self.parse_skc(skc)
        if not parsed:
            result['message'] = self._last_parse_error or "SKC 格式错误"
            return result

        new_style, color_code = parsed
        result['new_style'] = new_style
        result['color_code'] = color_code

        # 2. 查找老款号
        old_style = self.new_to_old_mapping.get(new_style)

        if not old_style:
            result['message'] = f"未找到新款号 {new_style} 对应的老款号"
            return result

        result['old_style'] = old_style

        # 3. 在店铺 SQLite 索引中查找老款号“同颜色”的所有尺码与源 SKU
        conn = self._connect_db(store_prefix)
        try:
            rows = conn.execute(
                """
                SELECT sku, size, suffix, source_file
                FROM ep_sku_index
                WHERE store_prefix = ?
                  AND style = ?
                  AND substr(sku, 8, 2) = ?
                ORDER BY CAST(size AS INTEGER), sku
                """,
                (store_prefix, old_style, color_code)
            ).fetchall()
        finally:
            conn.close()

        if not rows:
            result['message'] = (
                f"在 {store_prefix} 店铺聚合表中未找到老款号 {old_style} 颜色 {color_code} 的数据"
            )
            return result

        # 5. 构建结果
        normalized_items: List[Dict[str, str]] = []
        seen = set()
        for source_sku, size, suffix, source_file in rows:
            size_str = str(size)
            suffix_str = self._normalize_suffix(str(suffix or ""))
            if not str(suffix or "").strip():
                parent_sku = self._read_source_row_field(
                    source_file=str(source_file),
                    target_sku=str(source_sku),
                    field_names=[
                        "Parent SKU",
                        "child_parent_sku_relationship[marketplace_id=ATVPDKIKX0DER]#1.parent_sku",
                    ],
                )
                suffix_from_parent = self._extract_suffix_from_parent_sku(parent_sku)
                suffix_str = suffix_from_parent or self._generate_suffix_for_store(store_prefix, size_str)
            key = (size_str, suffix_str)
            if key in seen:
                continue
            seen.add(key)
            normalized_items.append({
                'size': size_str,
                'suffix': suffix_str,
                'sku': f"{new_style}{color_code}{size_str}{suffix_str}"  # 使用新款号构建 SKU
            })

        result['sizes'] = normalized_items

        result['success'] = True
        result['message'] = f"找到 {len(result['sizes'])} 个尺码"

        return result

    def get_sku_data_from_ep(
        self,
        old_style: str,
        color_code: str,
        size: str,
        suffix: str,
        template_type: str = "EPUS",
    ) -> Optional[List]:
        """从店铺数据中获取指定 SKU 的完整行数据

        Args:
            old_style: 老款号（7位）
            color_code: 颜色代码（2位）
            size: 尺码（2位）
            suffix: 后缀

        Returns:
            完整行数据列表，或 None
        """
        # 构建老款号的 SKU
        target_sku = f"{old_style}{color_code}{size}{suffix}"
        store_prefix = self._normalize_store_prefix(self.get_store_prefix(template_type))
        self._ensure_store_index(store_prefix)
        conn = self._connect_db(store_prefix)
        try:
            row = conn.execute(
                "SELECT source_file FROM ep_sku_index WHERE store_prefix = ? AND style = ? AND sku = ? LIMIT 1",
                (store_prefix, old_style, target_sku)
            ).fetchone()
        finally:
            conn.close()
        if not row:
            return None
        return self._load_row_data_by_sku(str(row[0]), target_sku)

    def _load_row_data_by_sku(self, ep_file: str, target_sku: str) -> Optional[List[Any]]:
        """按需从 EP 文件加载指定 SKU 的完整行数据"""
        file_path = UPLOADS_DIR / ep_file
        if not file_path.exists():
            return None

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb['Template']
            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) > 2 and row[2]:
                    sku = str(row[2]).strip()
                    if sku == target_sku:
                        wb.close()
                        return list(row)
            wb.close()
        except Exception as e:
            print(f"按需加载 SKU 行失败 {target_sku}: {e}")

        return None

    def _find_source_files(self, store_prefix: str, old_style: str, color_code: str) -> List[str]:
        normalized_store = self._normalize_store_prefix(store_prefix)
        self._ensure_store_index(normalized_store)
        conn = self._connect_db(normalized_store)
        try:
            rows = conn.execute(
                """
                SELECT DISTINCT source_file
                FROM ep_sku_index
                WHERE store_prefix = ?
                  AND style = ?
                  AND substr(sku, 8, 2) = ?
                ORDER BY source_file
                """,
                (normalized_store, old_style, color_code),
            ).fetchall()
        finally:
            conn.close()
        return [str(row[0]) for row in rows if row and row[0]]

    def process_skc(self, skc: str, store_prefix: str = "EP") -> Dict[str, Any]:
        template_type = self.get_template_type(store_prefix)
        query_result = self.find_sizes_for_skc(skc=skc, template_type=template_type)

        result: Dict[str, Any] = {
            "success": bool(query_result.get("success")),
            "skc": str(query_result.get("skc", skc)).strip().upper(),
            "old_style": str(query_result.get("old_style", "")).strip().upper(),
            "sizes": [
                str(item.get("size", "")).strip()
                for item in query_result.get("sizes", [])
                if str(item.get("size", "")).strip()
            ],
            "source_files": [],
            "message": str(query_result.get("message", "")).strip(),
        }

        if result["success"] and result["old_style"]:
            result["source_files"] = self._find_source_files(
                store_prefix=store_prefix,
                old_style=result["old_style"],
                color_code=str(query_result.get("color_code", "")).strip().upper(),
            )

        return result


# 全局单例
follow_sell_processor = FollowSellProcessor()
