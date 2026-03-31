"""
跟卖导出结果校验
"""
from __future__ import annotations

import asyncio
import json
import os
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Optional, Sequence

import openpyxl
from openpyxl.utils import get_column_letter

try:
    from anthropic import Anthropic
except ImportError:  # pragma: no cover - optional dependency
    Anthropic = None


ValidationItem = Dict[str, Any]
ValidationResult = Dict[str, Any]

TEMPLATE_SHEET_NAME = "Template"
HEADER_ROW = 2
MAPPING_VALUE_ROWS = (4, 5)
DATA_START_ROW = 4
LLM_PRIMARY_RANGE = range(80, 131)
LLM_SECONDARY_RANGE = range(268, 301)
MAPPING_REQUIRED_FIELDS = {}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _canonical_field_key(value: Any) -> str:
    return _normalize_header(value).replace(" ", "_")


def _make_issue(
    *,
    col: int,
    field: str,
    rule: str,
    detail: str,
) -> ValidationItem:
    return {
        "col": col,
        "col_letter": get_column_letter(col) if col and col > 0 else "",
        "field": field,
        "rule": rule,
        "detail": detail,
    }


def _build_header_maps(ws) -> tuple[Dict[str, List[int]], Dict[int, str]]:
    header_map: Dict[str, List[int]] = defaultdict(list)
    original_headers: Dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=HEADER_ROW, column=col).value
        header_text = str(header).strip() if header is not None else ""
        original_headers[col] = header_text
        if header_text:
            header_map[_normalize_header(header_text)].append(col)
    return dict(header_map), original_headers


def _find_columns(header_map: Dict[str, List[int]], *names: str) -> List[int]:
    found: List[int] = []
    for name in names:
        found.extend(header_map.get(_normalize_header(name), []))
    return list(dict.fromkeys(found))


def _first_non_empty_value(ws, row: int, columns: Sequence[int]) -> Any:
    for col in columns:
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def _is_non_empty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _is_excel_date(value: Any) -> bool:
    if isinstance(value, datetime):
        return True
    if isinstance(value, date):
        return True
    return False


def _to_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return Decimal(text.replace(",", ""))
    except InvalidOperation:
        return None


def _iter_data_rows(ws) -> Iterable[int]:
    for row in range(DATA_START_ROW, ws.max_row + 1):
        row_has_value = False
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if _is_non_empty(value):
                row_has_value = True
                break
        if row_has_value:
            yield row


def _collect_mapping_coverage(
    ws,
    template_type: str,
    header_map: Dict[str, List[int]],
    original_headers: Dict[int, str],
) -> List[ValidationItem]:
    errors: List[ValidationItem] = []
    required_fields = MAPPING_REQUIRED_FIELDS.get(str(template_type or "").strip(), [])
    if not required_fields:
        return errors

    alias_map = {
        "material_type": ("material type",),
        "colour_map": ("colour map", "color map"),
        "release_date": ("release date",),
        "restock_date": ("restock date",),
        "product_tax_code": ("product tax code",),
    }
    for field_key in required_fields:
        columns = _find_columns(header_map, field_key, *alias_map.get(field_key, ()))
        if not columns:
            errors.append(_make_issue(
                col=0,
                field=field_key,
                rule="mapping_coverage",
                detail=f"模板中未找到字段列: {field_key}",
            ))
            continue
        has_value = False
        for row in MAPPING_VALUE_ROWS:
            if _is_non_empty(_first_non_empty_value(ws, row, columns)):
                has_value = True
                break
        if has_value:
            continue
        col = columns[0]
        errors.append(_make_issue(
            col=col,
            field=original_headers.get(col) or field_key,
            rule="mapping_coverage",
            detail="第4或第5行映射值为空",
        ))
    return errors


def _collect_type_errors(ws, header_map: Dict[str, List[int]], original_headers: Dict[int, str]) -> List[ValidationItem]:
    errors: List[ValidationItem] = []
    date_columns = _find_columns(header_map, "Release Date", "Restock Date")
    numeric_columns = _find_columns(header_map, "Your Price", "Business Price")

    for row in _iter_data_rows(ws):
        for col in date_columns:
            value = ws.cell(row=row, column=col).value
            if not _is_non_empty(value):
                continue
            if _is_excel_date(value):
                continue
            errors.append(_make_issue(
                col=col,
                field=original_headers.get(col) or "",
                rule="type_check",
                detail=f"第{row}行应为日期，实际为 {type(value).__name__}: {value}",
            ))
        for col in numeric_columns:
            value = ws.cell(row=row, column=col).value
            if not _is_non_empty(value):
                continue
            if _to_decimal(value) is not None:
                continue
            errors.append(_make_issue(
                col=col,
                field=original_headers.get(col) or "",
                rule="type_check",
                detail=f"第{row}行应为数字，实际值: {value}",
            ))
    return errors


def _collect_price_errors(ws, header_map: Dict[str, List[int]], original_headers: Dict[int, str]) -> List[ValidationItem]:
    errors: List[ValidationItem] = []
    your_price_cols = _find_columns(header_map, "Your Price")
    business_price_cols = _find_columns(header_map, "Business Price")
    quantity_price_cols = []
    for index in range(1, 6):
        quantity_price_cols.extend(_find_columns(header_map, f"Quantity Price {index}"))
    if not quantity_price_cols:
        quantity_price_cols.extend([
            col for key, cols in header_map.items()
            if key.startswith("quantity price") and key != _normalize_header("Quantity Price Type")
            for col in cols
        ])
        quantity_price_cols = list(dict.fromkeys(quantity_price_cols))

    for row in _iter_data_rows(ws):
        your_price = _to_decimal(_first_non_empty_value(ws, row, your_price_cols))
        business_price = _to_decimal(_first_non_empty_value(ws, row, business_price_cols))
        if your_price is not None and business_price is not None:
            expected = your_price - Decimal("1")
            if abs(business_price - expected) > Decimal("0.1"):
                col = business_price_cols[0] if business_price_cols else (your_price_cols[0] if your_price_cols else 0)
                field = original_headers.get(col) or "Business Price"
                errors.append(_make_issue(
                    col=col,
                    field=field,
                    rule="price_relation",
                    detail=f"第{row}行 Business Price={business_price}，期望约等于 Your Price-1={expected}",
                ))

        ladder_prices: List[tuple[int, Decimal]] = []
        for col in quantity_price_cols:
            price = _to_decimal(ws.cell(row=row, column=col).value)
            if price is None:
                continue
            ladder_prices.append((col, price))
        for idx in range(1, len(ladder_prices)):
            prev_col, prev_price = ladder_prices[idx - 1]
            current_col, current_price = ladder_prices[idx]
            if current_price >= prev_price:
                errors.append(_make_issue(
                    col=current_col,
                    field=original_headers.get(current_col) or "Quantity Price",
                    rule="tiered_price_desc",
                    detail=f"第{row}行阶梯价未递减: {get_column_letter(prev_col)}={prev_price}, {get_column_letter(current_col)}={current_price}",
                ))
                break
    return errors


def _collect_parent_sku_warnings(ws, header_map: Dict[str, List[int]], original_headers: Dict[int, str]) -> List[ValidationItem]:
    warnings: List[ValidationItem] = []
    parent_sku_cols = _find_columns(header_map, "Parent SKU", "parent_sku")
    material_type_cols = _find_columns(header_map, "Material Type", "material_type")
    if not parent_sku_cols or not material_type_cols:
        return warnings

    grouped_materials: Dict[str, set[str]] = defaultdict(set)
    for row in _iter_data_rows(ws):
        parent_sku_value = _first_non_empty_value(ws, row, parent_sku_cols)
        material_value = _first_non_empty_value(ws, row, material_type_cols)
        if not _is_non_empty(parent_sku_value) or not _is_non_empty(material_value):
            continue
        parent_prefix = str(parent_sku_value).strip()[:7]
        material_text = str(material_value).strip()
        if parent_prefix and material_text:
            grouped_materials[parent_prefix].add(material_text)

    material_col = material_type_cols[0]
    material_field = original_headers.get(material_col) or "Material Type"
    for parent_prefix, materials in sorted(grouped_materials.items()):
        if len(materials) <= 1:
            continue
        warnings.append(_make_issue(
            col=material_col,
            field=material_field,
            rule="parent_material_consistency",
            detail=f"parent_sku 前7位 {parent_prefix} 存在多个 material_type: {', '.join(sorted(materials))}",
        ))
    return warnings


def _summarize_validation(errors: List[ValidationItem], warnings: List[ValidationItem]) -> str:
    if not errors and not warnings:
        return "校验通过，无错误或警告"
    return f"校验完成：{len(errors)} 个错误，{len(warnings)} 个警告"


def validate_output(wb_path: str, template_type: str) -> ValidationResult:
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    try:
        ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.active
        header_map, original_headers = _build_header_maps(ws)

        errors: List[ValidationItem] = []
        warnings: List[ValidationItem] = []

        errors.extend(_collect_mapping_coverage(ws, template_type, header_map, original_headers))
        errors.extend(_collect_type_errors(ws, header_map, original_headers))
        errors.extend(_collect_price_errors(ws, header_map, original_headers))
        warnings.extend(_collect_parent_sku_warnings(ws, header_map, original_headers))

        return {
            "passed": not errors,
            "warnings": warnings,
            "errors": errors,
            "summary": _summarize_validation(errors, warnings),
        }
    finally:
        wb.close()


async def _validate_with_llm_async(wb_path: str, template_type: str) -> ValidationResult:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key or Anthropic is None:
        return {}

    wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    try:
        ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.active
        headers = [ws.cell(row=HEADER_ROW, column=col).value for col in range(1, ws.max_column + 1)]
        header_map, _original_headers = _build_header_maps(ws)
        parent_sku_cols = _find_columns(header_map, "Parent SKU", "parent_sku")

        sampled_rows: Dict[str, Dict[str, Any]] = {}
        for row in _iter_data_rows(ws):
            parent_value = _first_non_empty_value(ws, row, parent_sku_cols) if parent_sku_cols else None
            parent_sku = str(parent_value).strip()[:7] if _is_non_empty(parent_value) else None
            key = parent_sku or f"row-{row}"
            if key in sampled_rows:
                continue
            cells: List[Dict[str, Any]] = []
            for col in list(LLM_PRIMARY_RANGE) + list(LLM_SECONDARY_RANGE):
                if col > ws.max_column:
                    continue
                cells.append({
                    "col": col,
                    "col_letter": get_column_letter(col),
                    "field": str(headers[col - 1] or ""),
                    "value": ws.cell(row=row, column=col).value,
                })
            sampled_rows[key] = {
                "row": row,
                "parent_sku": parent_sku or "",
                "cells": cells,
            }

        if not sampled_rows:
            return {}

        client = Anthropic(api_key=api_key)
        prompt = (
            "你是亚马逊模板质检助手。请检查以下 Excel 抽样行是否存在明显的语义问题，"
            "只返回 JSON 对象，格式为 "
            "{\"passed\": bool, \"warnings\": list, \"errors\": list, \"summary\": str}。"
            "warning/error 项格式必须是 "
            "{\"col\": int, \"col_letter\": str, \"field\": str, \"rule\": str, \"detail\": str}。"
            "rule 请使用 llm_semantic。"
            f"模板类型: {template_type}\n"
            f"抽样数据: {json.dumps(list(sampled_rows.values()), ensure_ascii=False, default=str)}"
        )
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1200,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
        )
        text_parts = [block.text for block in response.content if getattr(block, "type", "") == "text"]
        if not text_parts:
            return {}
        payload = json.loads("".join(text_parts).strip())
        if not isinstance(payload, dict):
            return {}
        return {
            "passed": bool(payload.get("passed", True)),
            "warnings": payload.get("warnings") or [],
            "errors": payload.get("errors") or [],
            "summary": str(payload.get("summary") or ""),
        }
    except Exception:
        return {}
    finally:
        wb.close()


def validate_with_llm(wb_path: str, template_type: str) -> ValidationResult:
    return asyncio.run(_validate_with_llm_async(wb_path, template_type))
