#!/usr/bin/env python3
"""
测试分店铺索引隔离：验证 ES01819 + NT + 04-28 在三个店铺的处理结果
"""
import sys
sys.path.insert(0, 'backend')

from pathlib import Path
from app.core.excel_processor import excel_processor
import openpyxl

# 生成测试 SKU
def generate_skus():
    skus = []
    for size in ['04', '06', '08', '10', '12', '14', '16', '18', '20', '22', '24', '26', '28']:
        skus.append(f"ES01819NT{size}-USA")
    return skus

# 三个店铺配置
STORES = [
    {'name': 'EP', 'template_type': 'EPUS', 'filenames': ['EP-0.xlsm', 'EP-1.xlsm', 'EP-2.xlsm']},
    {'name': 'DM', 'template_type': 'DaMaUS', 'filenames': ['DA-0.xlsm']},
    {'name': 'PZ', 'template_type': 'PZUS', 'filenames': ['PZ-0.xlsm', 'PZ-1.xlsm']},
]

def test_store(store_config):
    """测试单个店铺"""
    print(f"\n{'='*80}")
    print(f"测试店铺: {store_config['name']} ({store_config['template_type']})")
    print(f"{'='*80}")

    try:
        # 调用处理函数
        output_filename, processed_count = excel_processor.process_excel(
            template_type=store_config['template_type'],
            filenames=store_config['filenames'],
            selected_prefixes=['ES01819'],
            generated_skus=generate_skus(),
            processing_mode='add-color',
            progress_callback=lambda msg: print(f"  {msg}") if not msg.startswith('[DEBUG]') else None
        )

        print(f"\n✓ 处理完成: {output_filename}, 处理行数: {processed_count}")

        # 读取输出文件
        output_path = Path('backend/results') / output_filename
        if not output_path.exists():
            output_path = Path('backend/uploads') / output_filename

        if not output_path.exists():
            print(f"✗ 输出文件不存在: {output_filename}")
            return None

        wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
        ws = wb['Template'] if 'Template' in wb.sheetnames else wb.active

        # 读取表头
        header = [ws.cell(row=2, column=c).value for c in range(1, 10)]
        sku_col = next((i+1 for i, v in enumerate(header) if str(v).strip().lower() == 'seller sku'), 2)
        brand_col = next((i+1 for i, v in enumerate(header) if str(v).strip().lower() == 'brand name'), 3)
        name_col = next((i+1 for i, v in enumerate(header) if str(v).strip().lower() == 'product name'), 5)

        # 提取数据
        results = []
        for row in range(4, min(ws.max_row + 1, 50)):
            sku = ws.cell(row, sku_col).value
            if not sku or not str(sku).startswith('ES01819NT'):
                continue
            brand = ws.cell(row, brand_col).value
            name = ws.cell(row, name_col).value
            results.append({
                'sku': str(sku).strip(),
                'brand': str(brand or '').strip(),
                'name': str(name or '').strip()[:80]
            })

        wb.close()

        # 打印结果
        print(f"\n提取到 {len(results)} 行数据:")
        print(f"{'SKU':<20} {'Brand':<25} {'Product Name':<80}")
        print('-' * 125)
        for r in results[:5]:  # 只显示前5行
            print(f"{r['sku']:<20} {r['brand']:<25} {r['name']:<80}")
        if len(results) > 5:
            print(f"... (还有 {len(results) - 5} 行)")

        return results

    except Exception as e:
        print(f"✗ 处理失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    print("开始测试分店铺索引隔离...")

    all_results = {}
    for store in STORES:
        results = test_store(store)
        if results:
            all_results[store['name']] = results

    # 对比结果
    print(f"\n\n{'='*80}")
    print("对比分析")
    print(f"{'='*80}")

    if len(all_results) < 2:
        print("✗ 测试数据不足，无法对比")
        return

    # 检查品牌是否一致
    brands = {store: set(r['brand'] for r in results) for store, results in all_results.items()}
    print(f"\n各店铺品牌:")
    for store, brand_set in brands.items():
        print(f"  {store}: {brand_set}")

    # 检查是否有跨店铺串值
    if len(set(tuple(sorted(b)) for b in brands.values())) == 1:
        print("\n⚠️  警告: 所有店铺品牌完全一致，可能存在跨店铺串值！")
    else:
        print("\n✓ 各店铺品牌不同，索引隔离正常")

    # 检查产品名称前缀
    print(f"\n各店铺产品名称示例:")
    for store, results in all_results.items():
        if results:
            print(f"  {store}: {results[0]['name'][:60]}...")

if __name__ == '__main__':
    main()
